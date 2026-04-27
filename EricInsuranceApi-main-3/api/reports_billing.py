import csv
import io
from io import BytesIO
from sqlalchemy import bindparam, text
from constants import SORT_DIRECTIONS, ParticipantQueryFields
from models import User
from flask import Blueprint, Response, current_app, request, jsonify
from sqlalchemy.exc import SQLAlchemyError
from datetime import datetime
from db import Session as DBSession
from utils import build_where_clause, login_required
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.colors import black, white, grey, lightgrey
from reportlab.platypus import HRFlowable
from flask import send_file
import json
import os
import fitz

reports_billing_bp = Blueprint('reports_billing', __name__)


@reports_billing_bp.route('/', methods=['GET'])
@login_required
def reports_billing(user):

    search_by = request.args.get('search_by', 'Participant`')
    search_by_value = request.args.get('search_by_value', '')
    search_by_group = request.args.get('search_by_group', '')
    search_by_prem_due_date = request.args.get('search_by_prem_due_date', '')
    page = request.args.get('page', 1)
    page_size = request.args.get('page_size', 10)
    sort_by = request.args.get('sort_by', 'Prem_Due_Date')
    sort_order = request.args.get('sort_order', 'descend')

    if search_by not in ('Participant', 'Participant_SSN', 'Participant_Address', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    if isinstance(page, str):
        page = int(page)
    if isinstance(page_size, str):
        page_size = int(page_size)

    sql_params = {
        'offset': (page - 1) * page_size,
        'limit': page_size,
        'search_by_value': search_by_value,
    }

    if isinstance(search_by_value, str):
        search_by_clause = f'T1.[{search_by}] LIKE :search_by_value'
        if search_by_value == '':
            search_by_clause = f'{search_by_clause} OR T1.[{search_by}] IS NULL'
        sql_params['search_by_value'] = f'%{search_by_value}%'
    else:
        search_by_clause = f'T1.[{search_by}] = :search_by_value'

    if search_by_group != '':
        search_by_group_clause = f'T1.[Participant_Group_Name] = :search_by_group'
        sql_params['search_by_group'] = search_by_group
    else:
        search_by_group_clause = '1=1'

    if search_by_prem_due_date != '':
        search_by_prem_due_date_clause = f'T1.[Prem_Due_Date] = :search_by_prem_due_date'
        sql_params['search_by_prem_due_date'] = search_by_prem_due_date
    else:
        search_by_prem_due_date_clause = '1=1'

    if sort_by not in ('Participant', 'Participant_Group_Name', 'Prem_Due_Date'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT * FROM [dbo].[vw_Participant_Billing] AS T1
            WHERE ({search_by_clause})
            AND ({search_by_group_clause})
            AND ({search_by_prem_due_date_clause})
            ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            OFFSET :offset ROWS
            FETCH NEXT :limit ROWS ONLY;
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()
            data = [dict(zip(columns, row)) for row in rows]

            result_count = session.execute(text(f'''
            SELECT COUNT(1) as [Total] FROM [dbo].[vw_Participant_Billing] AS T1
            WHERE ({search_by_clause})
            AND ({search_by_group_clause})
            AND ({search_by_prem_due_date_clause});
            '''), sql_params)
            total = result_count.scalar()

            return jsonify({
                'columns': columns,
                'data': data,
                'total': total
            }), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@reports_billing_bp.route('/export_csv', methods=['GET'])
@login_required
def export_csv(user):

    search_by = request.args.get('search_by', 'Participant`')
    search_by_value = request.args.get('search_by_value', '')
    search_by_group = request.args.get('search_by_group', '')
    search_by_prem_due_date = request.args.get('search_by_prem_due_date', '')
    sort_by = request.args.get('sort_by', 'Participant')
    sort_order = request.args.get('sort_order', 'descend')

    if search_by not in ('Participant', 'Participant_SSN', 'Participant_Address', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    sql_params = {
        'search_by_value': search_by_value,
    }

    if search_by_prem_due_date == '' and search_by_group == '' and search_by_value == '':
        return jsonify(), '400 Parameters are not valid.'

    if isinstance(search_by_value, str):
        search_by_clause = f'T1.[{search_by}] LIKE :search_by_value'
        if search_by_value == '':
            search_by_clause = f'{search_by_clause} OR T1.[{search_by}] IS NULL'
        sql_params['search_by_value'] = f'%{search_by_value}%'
    else:
        search_by_clause = f'T1.[{search_by}] = :search_by_value'

    if search_by_group != '':
        search_by_group_clause = f'T1.[Participant_Group_Name] = :search_by_group'
        sql_params['search_by_group'] = search_by_group
    else:
        search_by_group_clause = '1=1'

    if search_by_prem_due_date != '':
        search_by_prem_due_date_clause = f'T1.[Prem_Due_Date] = :search_by_prem_due_date'
        sql_params['search_by_prem_due_date'] = search_by_prem_due_date
    else:
        search_by_prem_due_date_clause = '1=1'

    if sort_by == '':
        sort_by = 'Participant'
    if sort_order == '':
        sort_order = 'ascend'

    if sort_by not in ('Participant', 'Participant_Group_Name', 'Prem_Due_Date'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT * FROM [dbo].[vw_Participant_Billing] AS T1
            WHERE ({search_by_clause})
            AND ({search_by_group_clause})
            AND ({search_by_prem_due_date_clause})
            ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()

            # Write CSV to memory
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            for row in rows:
                writer.writerow(list(row))

            output.seek(0)
            csv_data = output.getvalue()
            output.close()

            # Prepare response
            response = Response(
                csv_data,
                mimetype='text/csv',
                headers={
                    "Content-Disposition": f"attachment; filename=billing_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
                }
            )
            return response
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'
        

@reports_billing_bp.route('/export_excel', methods=['GET'])
@login_required
def export_excel(user):
    search_by = request.args.get('search_by', 'Participant')
    search_by_value = request.args.get('search_by_value', '')
    search_by_group = request.args.get('search_by_group', '')
    search_by_prem_due_date = request.args.get('search_by_prem_due_date', '')
    sort_by = request.args.get('sort_by', 'Participant')
    sort_order = request.args.get('sort_order', 'descend')

    if search_by not in ('Participant', 'Participant_SSN', 'Participant_Address', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'

    sql_params = {'search_by_value': search_by_value}
    if search_by_prem_due_date == '' and search_by_group == '' and search_by_value == '':
        return jsonify(), '400 Parameters are not valid.'

    if isinstance(search_by_value, str):
        search_by_clause = f'T1.[{search_by}] LIKE :search_by_value'
        if search_by_value == '':
            search_by_clause += f' OR T1.[{search_by}] IS NULL'
        sql_params['search_by_value'] = f'%{search_by_value}%'
    else:
        search_by_clause = f'T1.[{search_by}] = :search_by_value'

    search_by_group_clause = '1=1'
    if search_by_group:
        search_by_group_clause = f'T1.[Participant_Group_Name] = :search_by_group'
        sql_params['search_by_group'] = search_by_group

    search_by_prem_due_date_clause = '1=1'
    if search_by_prem_due_date:
        search_by_prem_due_date_clause = f'T1.[Prem_Due_Date] = :search_by_prem_due_date'
        sql_params['search_by_prem_due_date'] = search_by_prem_due_date

    if sort_by not in ('Participant', 'Participant_Group_Name', 'Prem_Due_Date'):
        return jsonify(), '400 Parameters are not valid.'
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
                SELECT * FROM [dbo].[vw_Participant_Billing] AS T1
                WHERE ({search_by_clause})
                AND ({search_by_group_clause})
                AND ({search_by_prem_due_date_clause})
                ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            '''), sql_params)

            columns = list(result.keys())
            rows = result.fetchall()

            import openpyxl
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(columns)

            for row in rows:
                ws.append(list(row))

            for i, _ in enumerate(columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = 20

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    "Content-Disposition": f"attachment; filename=billing_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                }
            )

        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@reports_billing_bp.route('/billing_summary', methods=['GET'])
@login_required
def get_billing_summary(user):

    month = request.args.get('month', '')

    with DBSession() as session:
        try:
            # Build WHERE clause based on month parameter
            if month and month.strip():
                where_clause = "[Prem_Due_Date] = :month"
                sql_params = {'month': month}
            else:
                where_clause = "[Prem_Due_Date] > DATEADD(MONTH, -24, GETDATE())"
                sql_params = {}

            result = session.execute(text(f'''
            SELECT [Prem_Due_Date], [Prem_Group_PolicyNum], [Participant_Group_Name], COUNT(1) AS ParticipantsCount
            FROM [EKiser-Insurance-TablesSQL].[dbo].[vw_Participant_Billing]
            WHERE {where_clause}
            GROUP BY [Prem_Due_Date], [Prem_Group_PolicyNum], [Participant_Group_Name]
            ORDER BY [Prem_Due_Date] DESC, [Participant_Group_Name] ASC;
            '''), sql_params)
            
            columns = list(result.keys())
            rows = result.all()
            data = [dict(zip(columns, row)) for row in rows]

            return jsonify({
                'columns': columns,
                'data': data
            }), '200 OK'

        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@reports_billing_bp.route('/add_billing', methods=['POST'])
@login_required
def add_billing(user):
    data = request.json

    if 'month' not in data:
        return jsonify(), '400 Parameters are not valid.'
    
    month = data['month']

    if month == '':
        return jsonify(), '400 Parameters are not valid.'

    sql_params = {
        'month': month,
    }

    with DBSession() as session:
        try:
            check_result = session.execute(text('''
                SELECT COUNT(1) AS count
                FROM XParticipant_Billing
                WHERE Prem_Due_Date = :month
            '''), sql_params)
            
            existing_count = check_result.scalar()
            
            if existing_count > 0:
                return jsonify(), '400 Billing data already exists for this month.'
            
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'

    sql = text(f'''
    INSERT INTO XParticipant_Billing (
        Prem_Due_Date,
        Prem_Pkey,
        Prem_External_ID,
        Prem_Group_ID,
        Prem_Division,
        Prem_Group_PolicyNum,
        Prem_Baseline_Date,
        Prem_Term_Date,
        Prem_Status,
        Prem_BLife_EffectiveDate,
        Prem_BLife_EndDate,
        Prem_BLife_Amount,
        Prem_BLife_Count,
        Prem_BLife_Prem,
        Prem_BLife_ADD_Prem,
        Prem_BLife_Spouse_Amount,
        Prem_BLife_Child_Amount,
        Prem_BLife_Dep_Units,
        Prem_BLife_Dep_Prem,
        Prem_BLife_Total,
        Prem_VLife_EffectiveDate,
        Prem_VLife_EndDate,
        Prem_VLife_EE_Amount,
        Prem_VLife_EE_Count,
        Prem_VLife_EE_Prem,
        Prem_VLife_Spouse_Amount,
        Prem_VLife_Spouse_Count,
        Prem_VLife_Spouse_Prem,
        Prem_VLife_Child_Amount,
        Prem_VLife_Child_Count,
        Prem_VLife_Child_Prem,
        Prem_VLife_Total,
        [Prem_Total_Basic/VLife],
        Prem_VLife_Rate,
        Prem_Notes,
        Prem_Age,
        Prem_Age_Bracket,
        Prem_Weekly_VL_Deduction,
        Prem_Deduct_RequestSent,
        Prem_Payment_Type,
        Prem_Payment_Amount,
        Prem_Payment_Date
    )
    SELECT
        :month AS Prem_Due_Date,
        Participant_Pkey,
        Participant_External_ID,
        Participant_Group_ID,
        Participant_Division,
        Participant_Group_PolicyNum,
        Participant_Baseline_Date,
        Participant_TermDate,
        Participant_Status,
        Participant_BLife_EffectiveDate,
        Participant_BLife_EndDate,
        Participant_BLife_Amount,
        Participant_BLife_Count,
        Participant_BLife_Prem,
        Participant_BLife_ADD_Prem,
        Participant_BLife_Spouse_Amount,
        Participant_BLife_Child_Amount,
        Participant_BLife_Dep_Units,
        Participant_BLife_Dep_Prem,
        Participant_BLife_Total,
        Participant_VLife_EffectiveDate,
        Participant_VLife_EndDate,
        Participant_VLife_EE_Amount,
        Participant_VLife_EE_Count,
        Participant_VLife_EE_Prem,
        Participant_VLife_Spouse_Amount,
        Participant_VLife_Spouse_Count,
        Participant_VLife_Spouse_Prem,
        Participant_VLife_Child_Amount,
        Participant_VLife_Child_Count,
        Participant_VLife_Child_Prem,
        Participant_VLife_Total,
        [Participant_Total_Basic/VLife],
        Participant_VLife_Rate,
        Participant_Insurance_Notes,
        Participant_Age,
        Participant_Age_Bracket,
        Participant_Weekly_VL_Deduction,
        Participant_MailDate_DeductRequest,
        'E' AS Prem_Payment_Type,
        :month AS Prem_Payment_Date
    FROM XParticipant
    WHERE Participant_PKey > ' ' AND Participant_Status = 'ACTIVE'
    AND (
        (Participant_VLife_EE_Amount > 0 AND Participant_VLife_EndDate IS NULL) 
        OR (Participant_BLife_Amount > 0 AND Participant_BLife_EffectiveDate IS NOT NULL AND Participant_BLife_EndDate IS NULL)
    );
    ''')

    with DBSession() as session:
        try:
            session.execute(sql, sql_params)
            session.commit()
            return jsonify(), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@reports_billing_bp.route('/rollback_billing', methods=['POST'])
@login_required
def rollback_billing(user):
    data = request.json

    if 'month' not in data:
        return jsonify(), '400 Parameters are not valid.'
    
    month = data['month']

    if month == '':
        return jsonify(), '400 Parameters are not valid.'
    
    sql = text(f'''
    DELETE FROM XParticipant_Billing
    WHERE Prem_Due_Date = :month;
    ''')

    with DBSession() as session:
        try:
            session.execute(sql, {'month': month})
            session.commit()
            return jsonify(), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'



@reports_billing_bp.route('/by_pkeys', methods=['POST'])
@login_required
def participants_by_pkeys(user):
    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'

    pkeys = data.get('pkeys', [])
    if not pkeys:
        return jsonify(), '400 No pkeys provided.'

    pkeys = [str(pkey) for pkey in pkeys]
    pkeys_placeholders = [f':pkey{i}' for i in range(len(pkeys))]
    params = {f"pkey{i}": id_value for i, id_value in enumerate(pkeys)}

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT 
                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID
            FROM dbo.vw_participants T1
            LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
            WHERE T1.Participant_Pkey IN ({','.join(pkeys_placeholders)});
            '''), params)
            rows = result.all()
            columns = list(result.keys())
            data = [dict(zip(columns, row)) for row in rows]


            # load dependents
            dependent_sql = text('''
                SELECT [Depend_Pkey]
                    ,[Depend_AutoNum]
                    ,[Depend_ID]
                    ,[Depend_D_IDNum]
                    ,[Depend_Name]
                    ,[Depend_LastName]
                    ,[Depend_FirstName]
                    ,[Depend_MiddleName]
                    ,[Depend_SSN]
                    ,[Depend_Gender]
                    ,[Depend_TermDate]
                    ,[Depend_Relation]
                    ,[Depend_BirthDate]
                    ,[Depend_Coverage_Status]
                    ,[Depend_MED_Effective]
                    ,[Depend_MED_EndDate]
                    ,[Depend_DEN_Effective]
                    ,[Depend_DEN_EndDate]
                    ,[Depend_VIS_Effective]
                    ,[Depend_VIS_EndDate]
                    ,[Depend_ID1]
                    ,[Depend_ID2]
                    ,[Depend_Other_Insurance]
                    ,[Depend_PPO]
                    ,[Depend_InEligible_YN]
                    ,[Depend_WorkNotes]
                    ,[Depend_Marker]
                    ,[Depend_Change_Seq]
                    ,[Depend_PPO_Change_Seq]
                    ,[RowID]
                FROM dbo.XParticipant_Dependents
                WHERE Depend_Pkey IN :pkeys
                ORDER BY Depend_Pkey,Depend_AutoNum ASC
            ''').bindparams(bindparam('pkeys', expanding=True))

            dependent_results = session.execute(dependent_sql, {'pkeys': tuple(pkeys)}).mappings().fetchall()
            dependent_data = [dict(row) for row in dependent_results]
            dependent_data_dict = {}
            for dependent in dependent_data:
                if dependent['Depend_Pkey'] not in dependent_data_dict:
                    dependent_data_dict[dependent['Depend_Pkey']] = []
                dependent_data_dict[dependent['Depend_Pkey']].append(dependent)

            return jsonify({'participants': data, 'family_members': dependent_data_dict}), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'
        

@reports_billing_bp.route('/by_advanced_filter', methods=['POST'])
@login_required
def participants_by_advanced_filter(user):

    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'
    
    filter_criteria = data.get('filter_criteria')
    sort_by = data.get('sort_by', 'Name')
    sort_order = data.get('sort_order', 'ascend')

    if sort_by not in ('Name', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    sql_params = {}

    where_clause = build_where_clause(filter_criteria, sql_params)

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
                SELECT * FROM (
                    SELECT 
                    {ParticipantQueryFields},
                    T1.Terminated,
                    T1.RowID
                    FROM dbo.vw_participants T1
                        LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
                ) AS T1
                WHERE {where_clause}
                ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            '''), sql_params)
            rows = result.all()
            columns = list(result.keys())
            data = [dict(zip(columns, row)) for row in rows]

            pkeys = [row['Participant_Pkey'] for row in data if 'Participant_Pkey' in row]
            if not pkeys:
                return jsonify({'participants': [], 'family_members': {}}), '200 OK'

            dependent_sql = text(f'''
                SELECT [Depend_Pkey]
                    ,[Depend_AutoNum]
                    ,[Depend_ID]
                    ,[Depend_D_IDNum]
                    ,[Depend_Name]
                    ,[Depend_LastName]
                    ,[Depend_FirstName]
                    ,[Depend_MiddleName]
                    ,[Depend_SSN]
                    ,[Depend_Gender]
                    ,[Depend_TermDate]
                    ,[Depend_Relation]
                    ,[Depend_BirthDate]
                    ,[Depend_Coverage_Status]
                    ,[Depend_MED_Effective]
                    ,[Depend_MED_EndDate]
                    ,[Depend_DEN_Effective]
                    ,[Depend_DEN_EndDate]
                    ,[Depend_VIS_Effective]
                    ,[Depend_VIS_EndDate]
                    ,[Depend_ID1]
                    ,[Depend_ID2]
                    ,[Depend_Other_Insurance]
                    ,[Depend_PPO]
                    ,[Depend_InEligible_YN]
                    ,[Depend_WorkNotes]
                    ,[Depend_Marker]
                    ,[Depend_Change_Seq]
                    ,[Depend_PPO_Change_Seq]
                    ,[RowID]
                FROM dbo.XParticipant_Dependents
                WHERE Depend_Pkey IN (
                    SELECT Participant_Pkey FROM (
                        SELECT 
                        {ParticipantQueryFields},
                        T1.Terminated,
                        T1.RowID
                        FROM dbo.vw_participants T1
                            LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
                    ) AS T1
                    WHERE {where_clause}
                )
                ORDER BY Depend_Pkey,Depend_AutoNum ASC
            ''')

            dependent_results = session.execute(dependent_sql, sql_params).mappings().fetchall()
            dependent_data = [dict(row) for row in dependent_results]
            dependent_data_dict = {}
            for dependent in dependent_data:
                if dependent['Depend_Pkey'] not in dependent_data_dict:
                    dependent_data_dict[dependent['Depend_Pkey']] = []
                dependent_data_dict[dependent['Depend_Pkey']].append(dependent)

            return jsonify({'participants': data, 'family_members': dependent_data_dict}), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'

@reports_billing_bp.route('/gen_pdf_by_pkeys', methods=['POST'])
@login_required
def gen_pdf_by_pkeys(user):
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided.'}), 400

    pkeys = data.get('pkeys', [])
    if not pkeys:
        return jsonify({'error': 'No pkeys provided.'}), 400

    pkeys = [str(pkey) for pkey in pkeys]
    pkeys_placeholders = [f':pkey{i}' for i in range(len(pkeys))]
    params = {f"pkey{i}": id_value for i, id_value in enumerate(pkeys)}

    report_type = data.get('report_type', '')
    month_year = data.get('month_year')  # used only for report_type 110

    # Fetch PMT Weeks if report_type == '110'
    pmt_weeks = None
    if report_type == '110':
        if not month_year:
            return jsonify({'error': 'month_year is required for Payroll Deduction Analysis (report_type 110)'}), 400
        try:
            month, year = map(int, month_year.split("/"))
        except Exception:
            return jsonify({'error': 'Invalid month_year format. Use MM/YYYY'}), 400

        with DBSession() as session:
            result = session.execute(text('''
                SELECT SUBSTRING(Friday_PayDays, :month, 1) AS PMT_Weeks
                FROM TFridays
                WHERE Friday_Year = :year
            '''), {'month': month, 'year': year}).fetchone()
            if not result:
                return jsonify({'error': f'No entry in TFridays for {month_year}'}), 404
            pmt_weeks = int(result.PMT_Weeks)

    with DBSession() as session:
        try:
            # TODO: do sql query to get data related to report_type
            result = session.execute(text(f'''
                SELECT 
                    {ParticipantQueryFields},
                    T1.Terminated,
                    T1.RowID,
                    T1.Participant_Weekly_VL_Deduction
                FROM dbo.vw_participants T1
                LEFT JOIN Table_Rates ON (
                    T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID
                    AND T1.Participant_Baseline_Date = Table_Rates.Rate_Date
                )
                WHERE T1.Participant_Pkey IN ({','.join(pkeys_placeholders)});
            '''), params)
            rows = result.all()
            columns = list(result.keys())
            data = [dict(zip(columns, row)) for row in rows]

            # Fetch PPO data (latest per member ID)
            ppo_sql = text("""
                WITH RankedPPO AS (
                    SELECT
                        PPO_Member_Pkey,
                        PPO_Member_ID,
                        PPO_Network_01 AS PPO_Network1,
                        PPO_AccessPoint_01 AS PPO_AccessPoint1,
                        ROW_NUMBER() OVER (
                            PARTITION BY PPO_Member_Pkey
                            ORDER BY RowID ASC
                        ) AS rn
                    FROM dbo.XParticipant_PPO_Networks
                    WHERE PPO_Member_Pkey IN :pkeys
                )
                SELECT PPO_Member_Pkey, PPO_Member_ID, PPO_Network1, PPO_AccessPoint1
                FROM RankedPPO
                WHERE rn = 1;
            """).bindparams(bindparam("pkeys", expanding=True))


            ppo_results = session.execute(ppo_sql, {'pkeys': tuple(pkeys)}).mappings().fetchall()

            # Build lookup
            ppo_lookup = {
                str(row['PPO_Member_Pkey']).strip(): row
                for row in ppo_results
            }

            # Attach PPO to each participant
            for row in data:
                pkey = str(row.get('Participant_Pkey')).strip()
                ppo_data = ppo_lookup.get(pkey, {})
                row['PPO_Network1'] = ppo_data.get('PPO_Network1', '')
                row['PPO_AccessPoint1'] = ppo_data.get('PPO_AccessPoint1', '')
                row['PPO_Network_Display'] = f"{ppo_data.get('PPO_Network1', '')} Access Point: {ppo_data.get('PPO_AccessPoint1', '')}"

            # Load dependents
            dependent_sql = text('''
                SELECT [Depend_Pkey]
                    ,[Depend_AutoNum]
                    ,[Depend_ID]
                    ,[Depend_D_IDNum]
                    ,[Depend_Name]
                    ,[Depend_LastName]
                    ,[Depend_FirstName]
                    ,[Depend_MiddleName]
                    ,[Depend_SSN]
                    ,[Depend_Gender]
                    ,[Depend_TermDate]
                    ,[Depend_Relation]
                    ,[Depend_BirthDate]
                    ,[Depend_Coverage_Status]
                    ,[Depend_MED_Effective]
                    ,[Depend_MED_EndDate]
                    ,[Depend_DEN_Effective]
                    ,[Depend_DEN_EndDate]
                    ,[Depend_VIS_Effective]
                    ,[Depend_VIS_EndDate]
                    ,[Depend_ID1]
                    ,[Depend_ID2]
                    ,[Depend_Other_Insurance]
                    ,[Depend_PPO]
                    ,[Depend_InEligible_YN]
                    ,[Depend_WorkNotes]
                    ,[Depend_Marker]
                    ,[Depend_Change_Seq]
                    ,[Depend_PPO_Change_Seq]
                    ,[RowID]
                FROM dbo.XParticipant_Dependents
                WHERE Depend_Pkey IN :pkeys
                ORDER BY Depend_Pkey, Depend_AutoNum ASC
            ''').bindparams(bindparam('pkeys', expanding=True))

            dependent_results = session.execute(dependent_sql, {'pkeys': tuple(pkeys)}).mappings().fetchall()
            dependent_data = [dict(row) for row in dependent_results]
            dependent_data_dict = {}
            for dependent in dependent_data:
                dependent_data_dict.setdefault(dependent['Depend_Pkey'], []).append(dependent)

            kwargs = {
                'report_type': report_type,
                'participants': data,
                'family_members': dependent_data_dict
            }
            if report_type == '110':
                kwargs['pmt_weeks'] = pmt_weeks
                kwargs['month_year'] = month_year
                kwargs['session'] = session

            return generate_data_cleanup_pdf(**kwargs)

        except Exception as e:
            current_app.logerr(e)
            return jsonify({'error': 'An internal error has occurred.'}), 500


@reports_billing_bp.route('/gen_pdf_by_advanced_filter', methods=['POST'])
@login_required
def gen_pdf_by_advanced_filter(user):

    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'
    
    filter_criteria = data.get('filter_criteria')
    sort_by = data.get('sort_by', 'Name')
    sort_order = data.get('sort_order', 'ascend')

    report_type = data.get('report_type', '')

    if sort_by not in ('Name', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    sql_params = {}

    where_clause = build_where_clause(filter_criteria, sql_params)

    with DBSession() as session:
        try:
            # TODO: do sql query to get data related to report_type
            result = session.execute(text(f'''
                SELECT * FROM (
                    SELECT 
                    {ParticipantQueryFields},
                    T1.Terminated,
                    T1.RowID
                    FROM dbo.vw_participants T1
                        LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
                ) AS T1
                WHERE {where_clause}
                ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            '''), sql_params)
            rows = result.all()
            columns = list(result.keys())
            data = [dict(zip(columns, row)) for row in rows]

            dependent_sql = text(f'''
                SELECT [Depend_Pkey]
                    ,[Depend_AutoNum]
                    ,[Depend_ID]
                    ,[Depend_D_IDNum]
                    ,[Depend_Name]
                    ,[Depend_LastName]
                    ,[Depend_FirstName]
                    ,[Depend_MiddleName]
                    ,[Depend_SSN]
                    ,[Depend_Gender]
                    ,[Depend_TermDate]
                    ,[Depend_Relation]
                    ,[Depend_BirthDate]
                    ,[Depend_Coverage_Status]
                    ,[Depend_MED_Effective]
                    ,[Depend_MED_EndDate]
                    ,[Depend_DEN_Effective]
                    ,[Depend_DEN_EndDate]
                    ,[Depend_VIS_Effective]
                    ,[Depend_VIS_EndDate]
                    ,[Depend_ID1]
                    ,[Depend_ID2]
                    ,[Depend_Other_Insurance]
                    ,[Depend_PPO]
                    ,[Depend_InEligible_YN]
                    ,[Depend_WorkNotes]
                    ,[Depend_Marker]
                    ,[Depend_Change_Seq]
                    ,[Depend_PPO_Change_Seq]
                    ,[RowID]
                FROM dbo.XParticipant_Dependents
                WHERE Depend_Pkey IN (
                    SELECT Participant_Pkey FROM (
                        SELECT 
                        {ParticipantQueryFields},
                        T1.Terminated,
                        T1.RowID
                        FROM dbo.vw_participants T1
                            LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
                    ) AS T1
                    WHERE {where_clause}
                )
                ORDER BY Depend_Pkey,Depend_AutoNum ASC
            ''')

            dependent_results = session.execute(dependent_sql, sql_params).mappings().fetchall()
            dependent_data = [dict(row) for row in dependent_results]
            dependent_data_dict = {}
            for dependent in dependent_data:
                if dependent['Depend_Pkey'] not in dependent_data_dict:
                    dependent_data_dict[dependent['Depend_Pkey']] = []
                dependent_data_dict[dependent['Depend_Pkey']].append(dependent)

            # TODO: Generate PDF related to report_type
            return generate_data_cleanup_pdf(report_type=report_type, participants=data, family_members=dependent_data_dict)

        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


def generate_data_cleanup_pdf(**kwargs):
    """Generate PDF report using reportlab with proper page layout - 2 participants per page"""
    
    report_type = kwargs.get('report_type', '')
    if report_type == '100':
        participants = kwargs.get('participants')
        family_members = kwargs.get('family_members')
        return generate_report_100(participants)
    elif report_type == '105':
        participants = kwargs.get('participants')
        family_members = kwargs.get('family_members')
        return generate_report_105(participants)
    elif report_type == '110':
        participants = kwargs.get('participants')
        pmt_weeks = kwargs.get('pmt_weeks')
        month_year = kwargs.get('month_year')
        session = kwargs.get('session')
        return generate_payroll_deduction_analysis_110(participants, pmt_weeks, month_year, session)
    elif report_type == '115':
        participants = kwargs.get('participants')
        family_members = kwargs.get('family_members')
        return generate_age_reduction_115(participants)
    elif report_type == '120':
        participants = kwargs.get('participants')
        family_members = kwargs.get('family_members')
        return generate_summary_120_pdf(participants, family_members)
    elif report_type == '130':
        participants = kwargs.get('participants')
        family_members = kwargs.get('family_members')
        return generate_new_employee_130(participants, family_members)
    elif report_type == '150':
        participants = kwargs.get('participants')
        family_members = kwargs.get('family_members')
        return generate_report_150(participants, family_members)
    elif report_type == '140':
        participants = kwargs.get('participants')
        family_members = kwargs.get('family_members')
        return generate_report_140(participants, family_members)
    elif report_type == '200':
        participants = kwargs.get('participants')
        family_members = kwargs.get('family_members')
        return generate_report_200(participants, family_members)
    elif report_type == 'm03':
        participants = kwargs.get('participants')
        return generate_mailing_03(participants)
    elif report_type == 'm05':
        participants = kwargs.get('participants')
        return generate_mailing_05(participants)
    elif report_type == 'm06':
        participants = kwargs.get('participants')
        return generate_mailing_06(participants)
    elif report_type == 'm07':
        participants = kwargs.get('participants')
        return generate_mailing_07(participants)
    else:
        return jsonify({'error': 'Report type not supported.'}), 400
    

def generate_report_150(participants, family_members):
    # Create PDF buffer
    buffer = io.BytesIO()
    
    # Create PDF document in landscape A4 with page numbers
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=25*mm  # Increase bottom margin for page numbers
    )
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=12,
        fontName='Helvetica-Bold',
        spaceAfter=4,
        spaceBefore=0
    )
    
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=2,
        spaceBefore=0
    )
    
    # Build story (content)
    story = []
    
    # Helper functions
    def format_date(date_obj):
        if not date_obj:
            return 'N/A'
        try:
            if hasattr(date_obj, 'strftime'):
                return date_obj.strftime('%m/%d/%Y')
            return str(date_obj)
        except:
            return 'N/A'
    
    def format_ssn(ssn):
        if not ssn:
            return 'N/A'
        ssn_str = str(ssn).replace('-', '').replace('_', '')
        if len(ssn_str) >= 9:
            return f"{ssn_str[:3]}-{ssn_str[3:5]}-{ssn_str[5:]}"
        return ssn_str
    
    def calculate_age(birth_date):
        if not birth_date:
            return 'N/A'
        try:
            if hasattr(birth_date, 'year'):
                birth_year = birth_date.year
            else:
                birth_year = int(str(birth_date)[:4])
            current_year = datetime.now().year
            return str(current_year - birth_year)
        except:
            return 'N/A'
    
    def get_coverage_type_text(coverage_type):
        types = {
            'S': 'SINGLE',
            'F': 'FAMILY',
            'E': 'EMPLOYEE + SPOUSE',
            'C': 'EMPLOYEE + CHILDREN'
        }
        return types.get(coverage_type, coverage_type or 'N/A')
    
    def create_participant_content(participant, family_list):
        """Create content for a single participant"""
        content = []
        
        # Employee name - handle None values
        first_name = participant.get('Participant_FirstName') or ''
        middle_name = participant.get('Participant_MiddleInit') or ''
        last_name = participant.get('Participant_LastName') or ''
        employee_name = f"{first_name} {middle_name} {last_name}".strip()
        
        # 1. Title - First row with styled tag using table
        title_data = [['Employee', employee_name]]
        title_table = Table(title_data, colWidths=[1.2*inch, 6*inch])
        title_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),  # Left align the "Employee" label
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),  # Left align the employee name
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (0, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0, colors.white),  # No visible grid
            ('LEFTPADDING', (0, 0), (0, 0), 4),  # Reduce left padding to align to left edge
            ('RIGHTPADDING', (0, 0), (0, 0), 8),
            ('TOPPADDING', (0, 0), (0, 0), 4),
            ('BOTTOMPADDING', (0, 0), (0, 0), 4),
        ]))
        content.append(title_table)
        content.append(Spacer(1, 6))
        
        # 2. Basic Info - Left side (1/3 width)
        basic_info_data = [
            ['Employee', employee_name],
            ['Address', participant.get('Participant_Address', 'N/A')],
            ['City/State', f"{participant.get('Participant_City', '')}, {participant.get('Participant_State', '')} {participant.get('Participant_Zip', '')}"],
            ['Hired', format_date(participant.get('Participant_HireDate'))],
            ['Status', participant.get('Participant_Status', 'N/A')],
            ['Coverage', get_coverage_type_text(participant.get('Participant_Coverage_Type'))],
            ['SSN#', format_ssn(participant.get('Participant_SSN'))],
            ['Begin-Date', format_date(participant.get('Participant_BeginDate'))],
            ['Ending', format_date(participant.get('Participant_EndDate'))],
            ['Count', '4' if participant.get('Participant_Coverage_Type') == 'F' else '1']
        ]
        
        basic_info_table = Table(basic_info_data, colWidths=[1.0*inch, 1.5*inch])
        basic_info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        
        # 3. Coverage - Middle section (vertical layout)
        coverage_options = [
            ('Basic Life', participant.get('Participant_Coverage_BasicLife_YN')),
            ('Opt Out', participant.get('Participant_Coverage_OptOut_YN')),
            ('Med / RX', participant.get('Participant_Coverage_MedRX_YN')),
            ('Dental', participant.get('Participant_Coverage_Dental_YN')),
            ('Vision', participant.get('Participant_Coverage_Vision_YN')),
            ('COBRA', participant.get('Participant_Coverage_Cobra_YN')),
            ('VL-Waive', participant.get('Participant_Coverage_Waived_YN'))
        ]
        
        coverage_data = [['COVERAGE']]
        for option, checked in coverage_options:
            check_mark = '✓' if checked else '☐'
            coverage_data.append([f"{check_mark} {option}"])
        
        coverage_table = Table(coverage_data, colWidths=[1.8*inch])
        coverage_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        
        # 4. Family Members - Right side with Assumption note above
        # First create the assumption note
        assumption_data = [['Assumption CLEANUP NOTES']]
        assumption_table = Table(assumption_data, colWidths=[5.3*inch])
        assumption_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightcoral),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('MINIMUMHEIGHT', (0, 0), (-1, -1), 20),
        ]))
        
        # Then create the family members table
        family_data = [
            ['SEQ', 'RELATION', 'NAME', 'DOB', 'AGE', 'SSN', 'MEDICAL']
        ]
        
        # Add employee row
        family_data.append([
            '00',
            'Employee',
            employee_name,
            format_date(participant.get('Participant_BirthDate')),
            calculate_age(participant.get('Participant_BirthDate')),
            format_ssn(participant.get('Participant_SSN')),
            format_date(participant.get('Participant_MED_EffectiveDate'))
        ])
        
        # Add family members
        for k, member in enumerate(family_list):
            member_name = f"{member.get('Depend_FirstName', '')} {member.get('Depend_MiddleName', '')} {member.get('Depend_LastName', '')}".strip()
            family_data.append([
                f"{k+1:02d}",
                member.get('Depend_Relation', 'N/A'),
                member_name,
                format_date(member.get('Depend_BirthDate')),
                calculate_age(member.get('Depend_BirthDate')),
                format_ssn(member.get('Depend_SSN')),
                format_date(member.get('Depend_MED_Effective'))
            ])
        
        family_table = Table(family_data, colWidths=[0.4*inch, 0.7*inch, 1.5*inch, 0.7*inch, 0.4*inch, 0.8*inch, 0.8*inch])
        family_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        
        # Create a table to hold all three sections side by side with clear boundaries
        layout_data = [
            [basic_info_table, coverage_table, [assumption_table, family_table]]
        ]
        
        layout_table = Table(layout_data, colWidths=[2.5*inch, 1.8*inch, 5.3*inch])
        layout_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            # Add thick borders between sections
            ('GRID', (0, 0), (-1, -1), 2, colors.black),
            ('LINEBEFORE', (1, 0), (1, 0), 3, colors.darkblue),  # Thick line before coverage section
            ('LINEBEFORE', (2, 0), (2, 0), 3, colors.darkblue),  # Thick line before family section
            # Add background colors to distinguish sections
            ('BACKGROUND', (0, 0), (0, 0), colors.lightblue),
            ('BACKGROUND', (1, 0), (1, 0), colors.lightgreen),
            ('BACKGROUND', (2, 0), (2, 0), colors.lightyellow),
        ]))
        
        # Build content for this participant
        content.append(layout_table)
        content.append(Spacer(1, 8))
        
        return content
    
    # Process participants with proper page layout
    i = 0
    while i < len(participants):
        # Start with one participant per page
        page_content = []
        
        # Add first participant
        if i < len(participants):
            participant = participants[i]
            family_list = family_members.get(participant.get('Participant_Pkey'), [])
            participant_content = create_participant_content(participant, family_list)
            page_content.extend(participant_content)
            i += 1
        
        # Check if we can fit a second participant on the same page
        if i < len(participants):
            # Calculate estimated height for current page content
            current_height = len(page_content) * 20  # Rough estimate: each element ~20 points
            
            # Get second participant info
            second_participant = participants[i]
            second_family_list = family_members.get(second_participant.get('Participant_Pkey'), [])
            
            # Estimate height for second participant
            # Base height for participant content
            second_height = 200  # Base height for title, basic info, coverage, assumption
            
            # Add height for family members table
            family_rows = len(second_family_list) + 2  # +2 for header and employee row
            second_height += family_rows * 15  # ~15 points per row
            
            # Check if both participants can fit on one page
            # A4 landscape height is approximately 595 points, minus margins (~30 each side)
            available_height = 595 - 60  # ~535 points available
            
            if current_height + second_height <= available_height:
                # Add second participant to current page
                second_participant_content = create_participant_content(second_participant, second_family_list)
                page_content.extend(second_participant_content)
                i += 1
        
        # Add page content to story
        story.extend(page_content)
        
        # Add page break if there are more participants
        if i < len(participants):
            story.append(PageBreak())
    
    # Build PDF with page numbers
    def add_page_number(canvas, doc):
        page_num = canvas.getPageNumber()
        text = f"Page {page_num}"
        canvas.saveState()
        canvas.setFont('Helvetica', 10)
        canvas.drawCentredString(doc.pagesize[0]/2, 15*mm, text)
        canvas.restoreState()
    
    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    
    # Get PDF content
    buffer.seek(0)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = Response(
        pdf_content,
        mimetype='application/pdf',
        headers={
            "Content-Disposition": f"attachment; filename=data_cleanup_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        }
    )
    
    return response


def generate_report_140(participants, family_members):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer)

    full_story = []
    styles = getSampleStyleSheet()
    full_story.append(Paragraph(
        f"<b>Date Generated:</b> {datetime.now().strftime('%m/%d/%Y')}", styles["Normal"]
    ))

    for p in participants:
        dependents = family_members.get(p['Participant_Pkey'], [])
        full_story.extend(build_employer_report(p, dependents))  # append story for this participant

    if full_story and isinstance(full_story[-1], PageBreak):
        full_story.pop()

    doc.build(full_story, onFirstPage=add_page_number, onLaterPages=add_page_number)

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={
            "Content-Disposition": "attachment; filename=participant_summary.pdf"
        }
    )



def generate_report_200(participants, family_members):
    """
    Generate Mutual of Omaha Self-Bill Remittance Report (PR200) PDF.
    Creates a premium summary report with basic coverage and voluntary life tables.
    Receives pre-calculated data for different companies.
    """
    # The required modules are already imported at the top of the file, so no need to import them again here.

    def calculate_company_data(participants, family_members, company_name):
        """Calculate insurance data for a specific company based on 200report.md rules"""
        
        # Initialize data structures for each division (SHAR, SHIP, SHOP)

        divisions = []
        group_id = ''
        div_prefix_length = 0

        if company_name == "Shippers Rental Company":
            divisions = ['SHAR', 'SHIP', 'SHOP']
            group_id = '002370'
            div_prefix_length = 4
        elif company_name == "Sisbro, Inc.":
            divisions = ['SIS']
            group_id = '002620'
            div_prefix_length = 3

        basic_coverage = {div: {
            'employee_count': 0,
            'employee_volume': 0,
            'employee_premium': 0,
            'accidental_death': 0,
            'dependent_units': 0,
            'dependent_premium': 0,
            'total_premium': 0
        } for div in divisions}
        
        voluntary_life = {div: {
            'employee_count': 0,
            'employee_volume': 0,
            'employee_premium': 0,
            'spouse_count': 0,
            'spouse_volume': 0,
            'spouse_premium': 0,
            'dependent_count': 0,
            'dependent_volume': 0,
            'dependent_premium': 0,
            'total_premium': 0
        } for div in divisions}
        
        # Filter participants by company/group name
        company_participants = [p for p in participants if p.get('Participant_Group_ID') == group_id]
        
        for participant in company_participants:
            # Get division (first 4 characters of Participant_Division)
            division = participant.get('Participant_Division', '')
            division_prefix = division[:div_prefix_length] if division else ''
            
            # Only process if division matches divisions
            if division_prefix in divisions:
                # Basic Life Coverage calculations based on 200report.md rules
                bl_amount = participant.get('Participant_BLife_Amount') or 0
                bl_premium = participant.get('Participant_BLife_Prem') or 0
                add_premium = participant.get('Participant_BLife_ADD_Prem') or 0
                dep_units = participant.get('Participant_BLife_Dep_Units') or 0
                dep_premium = participant.get('Participant_BLife_Dep_Prem') or 0
                bl_total = participant.get('Participant_BLife_Total') or 0
                
                # Employee Count: Count if Participant_BLife_Amount > 0
                if bl_amount and float(bl_amount) > 0:
                    basic_coverage[division_prefix]['employee_count'] += 1
                
                # Employee Volume: Sum Participant_BLife_Amount if > 0
                if bl_amount and float(bl_amount) > 0:
                    basic_coverage[division_prefix]['employee_volume'] += float(bl_amount)
                
                # Employee Premium: Sum Participant_BLife_Prem if > 0
                if bl_premium and float(bl_premium) > 0:
                    basic_coverage[division_prefix]['employee_premium'] += float(bl_premium)
                
                # Accidental Death & Disability: Sum Participant_BLife_ADD_Prem if > 0
                if add_premium and float(add_premium) > 0:
                    basic_coverage[division_prefix]['accidental_death'] += float(add_premium)
                
                # Dependent Units: Sum Participant_BLife_Dep_Units if > 0
                if dep_units and int(dep_units) > 0:
                    basic_coverage[division_prefix]['dependent_units'] += int(dep_units)
                
                # Dependent Premium: Sum Participant_BLife_Dep_Prem if > 0
                if dep_premium and float(dep_premium) > 0:
                    basic_coverage[division_prefix]['dependent_premium'] += float(dep_premium)
                
                # BASIC COVERAGE TOTALS: Sum Participant_BLife_Total if > 0
                if bl_total and float(bl_total) > 0:
                    basic_coverage[division_prefix]['total_premium'] += float(bl_total)
                
                # Voluntary Life calculations
                vl_ee_amount = participant.get('Participant_VLife_EE_Amount') or 0
                vl_ee_premium = participant.get('Participant_VLife_EE_Prem') or 0
                vl_spouse_amount = participant.get('Participant_VLife_Spouse_Amount') or 0
                vl_spouse_premium = participant.get('Participant_VLife_Spouse_Prem') or 0
                vl_child_amount = participant.get('Participant_VLife_Child_Amount') or 0
                vl_child_premium = participant.get('Participant_VLife_Child_Prem') or 0
                
                # Employee voluntary life
                if vl_ee_amount and float(vl_ee_amount) > 0:
                    voluntary_life[division_prefix]['employee_count'] += 1
                    voluntary_life[division_prefix]['employee_volume'] += float(vl_ee_amount)
                    if vl_ee_premium and float(vl_ee_premium) > 0:
                        voluntary_life[division_prefix]['employee_premium'] += float(vl_ee_premium)
                
                # Spouse voluntary life
                if vl_spouse_amount and float(vl_spouse_amount) > 0:
                    voluntary_life[division_prefix]['spouse_count'] += 1
                    voluntary_life[division_prefix]['spouse_volume'] += float(vl_spouse_amount)
                    if vl_spouse_premium and float(vl_spouse_premium) > 0:
                        voluntary_life[division_prefix]['spouse_premium'] += float(vl_spouse_premium)
                
                # Child voluntary life
                if vl_child_amount and float(vl_child_amount) > 0:
                    voluntary_life[division_prefix]['dependent_count'] += 1
                    voluntary_life[division_prefix]['dependent_volume'] += float(vl_child_amount)
                    if vl_child_premium and float(vl_child_premium) > 0:
                        voluntary_life[division_prefix]['dependent_premium'] += float(vl_child_premium)
        
        # Calculate totals for each division
        for div in divisions:
            # Basic coverage totals (if not already calculated from Participant_BLife_Total)
            if basic_coverage[div]['total_premium'] == 0:
                basic_coverage[div]['total_premium'] = (
                    basic_coverage[div]['employee_premium'] + 
                    basic_coverage[div]['accidental_death'] + 
                    basic_coverage[div]['dependent_premium']
                )
            
            # Voluntary life totals
            voluntary_life[div]['total_premium'] = (
                voluntary_life[div]['employee_premium'] + 
                voluntary_life[div]['spouse_premium'] + 
                voluntary_life[div]['dependent_premium']
            )
        
        return {
            'basic_coverage': basic_coverage,
            'voluntary_life': voluntary_life,
            'participant_count': len(company_participants)
        }

    # Calculate Shippers Rental Company data
    shippers_rental_company_data = calculate_company_data(participants, family_members, "Shippers Rental Company")

    # Calculate Sisbro, Inc. data
    sisbro_inc_data = calculate_company_data(participants, family_members, "Sisbro, Inc.")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.75 * inch
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle('Title200', parent=styles['Title'], fontSize=16, alignment=1, spaceAfter=12, fontName='Helvetica-Bold')
    subtitle_style = ParagraphStyle('Subtitle200', parent=styles['Normal'], fontSize=12, alignment=1, spaceAfter=8, fontName='Helvetica-Bold')
    header_style = ParagraphStyle('Header200', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', alignment=1)
    normal_style = ParagraphStyle('Normal200', parent=styles['Normal'], fontSize=9, alignment=0)
    small_style = ParagraphStyle('Small200', parent=styles['Normal'], fontSize=8, alignment=0)
    
    def format_currency(amount):
        """Format amount as currency"""
        if amount is None or amount == 0:
            return "$0.00"
        try:
            return f"${float(amount):.2f}"
        except:
            return "$0.00"
    
    def format_count(count):
        """Format count as integer"""
        if count is None or count == 0:
            return "0"
        try:
            return str(int(float(count)))
        except:
            return "0"
    
    def format_volume(volume):
        """Format volume as currency"""
        if volume is None or volume == 0:
            return "$0"
        try:
            return f"${int(float(volume)):,}"
        except:
            return "$0"
    
    story = []
    
    # Title
    # Add a black background, white text title at the top of every page using onPage callback
    def draw_header(canvas, doc):
        canvas.saveState()
        canvas.setFillColorRGB(0, 0, 0)
        canvas.rect(0, doc.height + doc.topMargin + 0.25*inch, doc.width + doc.leftMargin + doc.rightMargin, 0.5*inch, fill=1, stroke=0)
        canvas.setFillColorRGB(1, 1, 1)
        canvas.setFont("Helvetica-Bold", 16)
        text_width = canvas.stringWidth("ERIC KISER INSURANCE", "Helvetica-Bold", 16)
        x = (doc.width + doc.leftMargin + doc.rightMargin - text_width) / 2
        y = doc.height + doc.topMargin + 0.4*inch
        canvas.drawString(x, y, "ERIC KISER INSURANCE")
        canvas.restoreState()
    
    story.append(Paragraph("Mutual of Omaha Self-Bill Remittance Report", title_style))
    story.append(Paragraph("Rept: PR200", small_style))
    story.append(Spacer(1, 0.2 * inch))
    
    # Shippers Rental Company data
    if shippers_rental_company_data:
        # Use a table to align the company name to the left and the participant count to the right on the same line
        company_name = Paragraph("Shippers Rental Company", header_style)
        participant_count = Paragraph(f'{shippers_rental_company_data.get("participant_count", 0)}-Participants', normal_style)
        header_table = Table(
            [[company_name, participant_count]],
            colWidths=[3.5*inch, 3.5*inch]
        )
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 0.1 * inch))
        
        # Get basic coverage data
        basic_data = shippers_rental_company_data.get('basic_coverage', {})
        voluntary_data = shippers_rental_company_data.get('voluntary_life', {})
        
        # Determine column headers based on data structure
        columns = ['SHAR', 'SHIP', 'SHOP', 'TOTAL']
        
        # Calculate totals for each division
        shar_basic = basic_data.get('SHAR', {})
        ship_basic = basic_data.get('SHIP', {})
        shop_basic = basic_data.get('SHOP', {})
        
        # BASIC COVERAGE Table
        basic_table_data = [
            ['BASIC COVERAGE'] + columns,
            ['Employee Count'] + [
                format_count(shar_basic.get('employee_count', 0)),
                format_count(ship_basic.get('employee_count', 0)),
                format_count(shop_basic.get('employee_count', 0)),
                format_count(shar_basic.get('employee_count', 0) + ship_basic.get('employee_count', 0) + shop_basic.get('employee_count', 0))
            ],
            ['Employee Volume'] + [
                format_volume(shar_basic.get('employee_volume', 0)),
                format_volume(ship_basic.get('employee_volume', 0)),
                format_volume(shop_basic.get('employee_volume', 0)),
                format_volume(shar_basic.get('employee_volume', 0) + ship_basic.get('employee_volume', 0) + shop_basic.get('employee_volume', 0))
            ],
            ['Employee Premium'] + [
                format_currency(shar_basic.get('employee_premium', 0)),
                format_currency(ship_basic.get('employee_premium', 0)),
                format_currency(shop_basic.get('employee_premium', 0)),
                format_currency(shar_basic.get('employee_premium', 0) + ship_basic.get('employee_premium', 0) + shop_basic.get('employee_premium', 0))
            ],
            ['Accidental Death & Disab'] + [
                format_currency(shar_basic.get('accidental_death', 0)),
                format_currency(ship_basic.get('accidental_death', 0)),
                format_currency(shop_basic.get('accidental_death', 0)),
                format_currency(shar_basic.get('accidental_death', 0) + ship_basic.get('accidental_death', 0) + shop_basic.get('accidental_death', 0))
            ],
            ['Dependent Units'] + [
                format_count(shar_basic.get('dependent_units', 0)),
                format_count(ship_basic.get('dependent_units', 0)),
                format_count(shop_basic.get('dependent_units', 0)),
                format_count(shar_basic.get('dependent_units', 0) + ship_basic.get('dependent_units', 0) + shop_basic.get('dependent_units', 0))
            ],
            ['Dependent Premium'] + [
                format_currency(shar_basic.get('dependent_premium', 0)),
                format_currency(ship_basic.get('dependent_premium', 0)),
                format_currency(shop_basic.get('dependent_premium', 0)),
                format_currency(shar_basic.get('dependent_premium', 0) + ship_basic.get('dependent_premium', 0) + shop_basic.get('dependent_premium', 0))
            ],
            ['BASIC COVERAGE TOTALS'] + [
                format_currency(shar_basic.get('total_premium', 0)),
                format_currency(ship_basic.get('total_premium', 0)),
                format_currency(shop_basic.get('total_premium', 0)),
                format_currency(shar_basic.get('total_premium', 0) + ship_basic.get('total_premium', 0) + shop_basic.get('total_premium', 0))
            ]
        ]
        
        basic_table = Table(basic_table_data, colWidths=[2.5*inch] + [1.1*inch]*4)
        basic_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        story.append(basic_table)
        story.append(Spacer(1, 0.15 * inch))
        
        # Calculate totals for voluntary life
        shar_vol = voluntary_data.get('SHAR', {})
        ship_vol = voluntary_data.get('SHIP', {})
        shop_vol = voluntary_data.get('SHOP', {})
        
        # VOLUNTARY LIFE Table
        voluntary_table_data = [
            ['VOLUNTARY LIFE'] + columns,
            ['Employee Count'] + [
                format_count(shar_vol.get('employee_count', 0)),
                format_count(ship_vol.get('employee_count', 0)),
                format_count(shop_vol.get('employee_count', 0)),
                format_count(shar_vol.get('employee_count', 0) + ship_vol.get('employee_count', 0) + shop_vol.get('employee_count', 0))
            ],
            ['Employee Volume'] + [
                format_volume(shar_vol.get('employee_volume', 0)),
                format_volume(ship_vol.get('employee_volume', 0)),
                format_volume(shop_vol.get('employee_volume', 0)),
                format_volume(shar_vol.get('employee_volume', 0) + ship_vol.get('employee_volume', 0) + shop_vol.get('employee_volume', 0))
            ],
            ['Employee Premium'] + [
                format_currency(shar_vol.get('employee_premium', 0)),
                format_currency(ship_vol.get('employee_premium', 0)),
                format_currency(shop_vol.get('employee_premium', 0)),
                format_currency(shar_vol.get('employee_premium', 0) + ship_vol.get('employee_premium', 0) + shop_vol.get('employee_premium', 0))
            ],
            # Add a blank row for spacing
            [],
            ['Spouse Count'] + [
                format_count(shar_vol.get('spouse_count', 0)),
                format_count(ship_vol.get('spouse_count', 0)),
                format_count(shop_vol.get('spouse_count', 0)),
                format_count(shar_vol.get('spouse_count', 0) + ship_vol.get('spouse_count', 0) + shop_vol.get('spouse_count', 0))
            ],
            ['Spouse Volume'] + [
                format_volume(shar_vol.get('spouse_volume', 0)),
                format_volume(ship_vol.get('spouse_volume', 0)),
                format_volume(shop_vol.get('spouse_volume', 0)),
                format_volume(shar_vol.get('spouse_volume', 0) + ship_vol.get('spouse_volume', 0) + shop_vol.get('spouse_volume', 0))
            ],
            ['Spouse Premium'] + [
                format_currency(shar_vol.get('spouse_premium', 0)),
                format_currency(ship_vol.get('spouse_premium', 0)),
                format_currency(shop_vol.get('spouse_premium', 0)),
                format_currency(shar_vol.get('spouse_premium', 0) + ship_vol.get('spouse_premium', 0) + shop_vol.get('spouse_premium', 0))
            ],
            [],
            ['Dependent Count'] + [
                format_count(shar_vol.get('dependent_count', 0)),
                format_count(ship_vol.get('dependent_count', 0)),
                format_count(shop_vol.get('dependent_count', 0)),
                format_count(shar_vol.get('dependent_count', 0) + ship_vol.get('dependent_count', 0) + shop_vol.get('dependent_count', 0))
            ],
            ['Dependent Volume'] + [
                format_volume(shar_vol.get('dependent_volume', 0)),
                format_volume(ship_vol.get('dependent_volume', 0)),
                format_volume(shop_vol.get('dependent_volume', 0)),
                format_volume(shar_vol.get('dependent_volume', 0) + ship_vol.get('dependent_volume', 0) + shop_vol.get('dependent_volume', 0))
            ],
            ['Dependent Premium'] + [
                format_currency(shar_vol.get('dependent_premium', 0)),
                format_currency(ship_vol.get('dependent_premium', 0)),
                format_currency(shop_vol.get('dependent_premium', 0)),
                format_currency(shar_vol.get('dependent_premium', 0) + ship_vol.get('dependent_premium', 0) + shop_vol.get('dependent_premium', 0))
            ],
            ['VOLUNTARY LIFE TOTALS'] + [
                format_currency(shar_vol.get('total_premium', 0)),
                format_currency(ship_vol.get('total_premium', 0)),
                format_currency(shop_vol.get('total_premium', 0)),
                format_currency(shar_vol.get('total_premium', 0) + ship_vol.get('total_premium', 0) + shop_vol.get('total_premium', 0))
            ]
        ]
        
        voluntary_table = Table(voluntary_table_data, colWidths=[2.5*inch] + [1.1*inch]*4)
        voluntary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        story.append(voluntary_table)
        story.append(Spacer(1, 0.1 * inch))
        
        # Overall Total
        total_basic = (shar_basic.get('total_premium', 0) + ship_basic.get('total_premium', 0) + shop_basic.get('total_premium', 0))
        total_vol = (shar_vol.get('total_premium', 0) + ship_vol.get('total_premium', 0) + shop_vol.get('total_premium', 0))
        total_basic_vol = total_basic + total_vol
        
        total_data = [
            ['TOTAL BASIC & VOL.LIFE'] + [format_currency(total_basic_vol)] * 4
        ]
        
        total_table = Table(total_data, colWidths=[2.5*inch] + [1.1*inch]*4)
        total_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        story.append(total_table)
        story.append(Spacer(1, 0.2 * inch))
    
    # Sisbro, Inc.
    if sisbro_inc_data:
        # Add a page break before Sisbro, Inc. section to ensure it starts on a new page
        story.append(PageBreak())
        # Use a table to align the company name to the left and the participant count to the right on the same line, similar to shippers_rental_company_data
        company_name = Paragraph("Sisbro, Inc.", header_style)
        participant_count = Paragraph(f'{sisbro_inc_data.get("participant_count", 0)}-Participants', normal_style)
        header_table = Table(
            [[company_name, participant_count]],
            colWidths=[3.5*inch, 3.5*inch]
        )
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 0.1 * inch))

        # Get basic coverage data
        basic_data = sisbro_inc_data.get('basic_coverage', {})
        voluntary_data = sisbro_inc_data.get('voluntary_life', {})

        # Sisbro division: SIS
        columns = ['SIS', 'TOTAL']

        sis_basic = basic_data.get('SIS', {})
        sis_vol = voluntary_data.get('SIS', {})

        # BASIC COVERAGE Table
        basic_table_data = [
            ['BASIC COVERAGE'] + columns,
            ['Employee Count'] + [
                format_count(sis_basic.get('employee_count', 0)),
                format_count(sis_basic.get('employee_count', 0))
            ],
            ['Employee Volume'] + [
                format_volume(sis_basic.get('employee_volume', 0)),
                format_volume(sis_basic.get('employee_volume', 0))
            ],
            ['Employee Premium'] + [
                format_currency(sis_basic.get('employee_premium', 0)),
                format_currency(sis_basic.get('employee_premium', 0))
            ],
            ['Accidental Death & Disab'] + [
                format_currency(sis_basic.get('accidental_death', 0)),
                format_currency(sis_basic.get('accidental_death', 0))
            ],
            ['Dependent Units'] + [
                format_count(sis_basic.get('dependent_units', 0)),
                format_count(sis_basic.get('dependent_units', 0))
            ],
            ['Dependent Premium'] + [
                format_currency(sis_basic.get('dependent_premium', 0)),
                format_currency(sis_basic.get('dependent_premium', 0))
            ],
            ['TOTAL BASIC'] + [
                format_currency(sis_basic.get('total_premium', 0)),
                format_currency(sis_basic.get('total_premium', 0))
            ]
        ]

        basic_table = Table(basic_table_data, colWidths=[2.5*inch] + [1.1*inch]*2)
        basic_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))

        story.append(basic_table)
        story.append(Spacer(1, 0.1 * inch))

        # VOLUNTARY LIFE Table
        voluntary_table_data = [
            ['VOLUNTARY LIFE'] + columns,
            ['Employee Count'] + [
                format_count(sis_vol.get('employee_count', 0)),
                format_count(sis_vol.get('employee_count', 0))
            ],
            ['Employee Volume'] + [
                format_volume(sis_vol.get('employee_volume', 0)),
                format_volume(sis_vol.get('employee_volume', 0))
            ],
            ['Employee Premium'] + [
                format_currency(sis_vol.get('employee_premium', 0)),
                format_currency(sis_vol.get('employee_premium', 0))
            ],
            [],
            ['Spouse Count'] + [
                format_count(sis_vol.get('spouse_count', 0)),
                format_count(sis_vol.get('spouse_count', 0))
            ],
            ['Spouse Volume'] + [
                format_volume(sis_vol.get('spouse_volume', 0)),
                format_volume(sis_vol.get('spouse_volume', 0))
            ],
            ['Spouse Premium'] + [
                format_currency(sis_vol.get('spouse_premium', 0)),
                format_currency(sis_vol.get('spouse_premium', 0))
            ],
            [],
            ['Dependent Count'] + [
                format_count(sis_vol.get('dependent_count', 0)),
                format_count(sis_vol.get('dependent_count', 0))
            ],
            ['Dependent Volume'] + [
                format_volume(sis_vol.get('dependent_volume', 0)),
                format_volume(sis_vol.get('dependent_volume', 0))
            ],
            ['Dependent Premium'] + [
                format_currency(sis_vol.get('dependent_premium', 0)),
                format_currency(sis_vol.get('dependent_premium', 0))
            ],
            ['TOTAL VOL.LIFE'] + [
                format_currency(sis_vol.get('total_premium', 0)),
                format_currency(sis_vol.get('total_premium', 0))
            ]
        ]

        voluntary_table = Table(voluntary_table_data, colWidths=[2.5*inch] + [1.1*inch]*2)
        voluntary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))

        story.append(voluntary_table)
        story.append(Spacer(1, 0.1 * inch))

        # Overall Total
        total_basic = sis_basic.get('total_premium', 0)
        total_vol = sis_vol.get('total_premium', 0)
        total_basic_vol = total_basic + total_vol

        total_data = [
            ['TOTAL BASIC & VOL.LIFE'] + [format_currency(total_basic_vol)] * 2
        ]

        total_table = Table(total_data, colWidths=[2.5*inch] + [1.1*inch]*2)
        total_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))

        story.append(total_table)
        story.append(Spacer(1, 0.2 * inch))

    # Footer function
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        now = datetime.now().strftime('Printed: %A %b-%d-%Y %I:%M %p')
        canvas.drawString(0.5*inch, 0.5*inch, now)
        canvas.restoreState()
    
    # Combined header and footer function
    def on_first_page(canvas, doc):
        draw_header(canvas, doc)
        add_footer(canvas, doc)
    
    def on_later_pages(canvas, doc):
        draw_header(canvas, doc)
        add_footer(canvas, doc)
    
    doc.build(story, onFirstPage=on_first_page, onLaterPages=on_later_pages)
    buffer.seek(0)
    
    from flask import Response
    return Response(
        buffer.read(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "inline; filename=PR200_premium_summary.pdf"}
    )


def generate_report_100(participants):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.75 * inch
    )

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    normal_style = styles['Normal']

    story = []

    # Title
    story.append(Paragraph("Master Listing", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 0.2 * inch))

    # Table headers
    table_data = [[
        'Name', 'Address', 'City/State', 'Hires', 'ID', 'Group', 'Hired',
        'Med Effective', 'Med End', 'Type', 'COBRA', 'Status'
    ]]
    # Populate rows
    for p in participants:
        name = f"{p.get('Participant_FirstName', '')} {p.get('Participant_LastName', '')}".strip()
        address = p.get('Participant_Address', '')
        if address and len(address) > 23:
            address = address[:19] + "…"
        city_state = f"{p.get('Participant_City', '')}, {p.get('Participant_State', '')} {p.get('Participant_Zip') or ''}".split('-')[0]
        if city_state and len(city_state) > 23:
            city_state = city_state[:19] + "…"
        ssn = str(p.get('Participant_SSN', ''))[-4:] if p.get('Participant_SSN') else ''
        group = p.get('Participant_Group_Name', '')
        hire = p.get('Participant_HireDate')
        med_start = p.get('Participant_MED_EffectiveDate')
        med_end = p.get('Participant_MED_EndDate')
        cov_type = p.get('Participant_Coverage_Type', '')
        cobra = 'Yes' if p.get('Participant_Coverage_Cobra_YN') else ''
        status = 'T' if p.get('Participant_Status') == 'TERM' else 'A' if p.get('Participant_Status') == 'ACTIVE' else p.get('Participant_Status', '')
        ssn_raw = p.get('Participant_SSN')
        hires = sum(1 for entry in participants if entry.get('Participant_SSN') == ssn_raw)

        row = [
            name,
            address,
            city_state,
            str(hires),
            ssn,
            group,
            hire.strftime('%m/%d/%Y') if hire else '',
            med_start.strftime('%m/%d/%Y') if med_start else '',
            med_end.strftime('%m/%d/%Y') if med_end else '',
            cov_type,
            cobra,
            status
        ]
        table_data.append(row)

    # Create and style table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (1, 0), (-1, -1), [colors.white, colors.whitesmoke]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))

    story.append(table)

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)

    buffer.seek(0)
    return Response(
        buffer.read(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "inline; filename=master_listing.pdf"}
    )


def generate_report_105(participants):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=20 * mm
    )

    styles = getSampleStyleSheet()
    story = []
    timestamp = datetime.now().strftime("%B %d, %Y %I:%M %p")

    story.append(Paragraph("Insurance Monthly", styles["Title"]))
    story.append(Paragraph(f"Generated on: {timestamp}", styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    # Table headers: two-level
    header1 = [
        '', '', '',
        'BASIC COVERAGE - Employee', '', '',
        'VOLUNTARY LIFE - Employee', '', '',
        'VOLUNTARY LIFE - Spouse', '',
        'VOLUNTARY LIFE - Child', '',
        '', '', ''
    ]
    header2 = [
        'Name', 'ID', 'Group',
        'Effective', 'Amt', 'Dep Units',
        'Effective', 'Amount', 'Prem',
        'Amount', 'Prem',
        'Amount', 'Prem',
        'VLife Total', 'Weekly Deduct', 'Termed'
    ]

    data = [header1, header2]

    # Populate rows
    for p in participants:
        row = [
            f"{p.get('Participant_FirstName', '')} {p.get('Participant_LastName', '')}",
            str(p.get('Participant_SSN'))[-4:] if p.get('Participant_SSN') else '',
            p.get('Participant_Group_Name', ''),
            p.get('Participant_BLife_EffectiveDate').strftime('%m/%d/%Y') if p.get('Participant_BLife_EffectiveDate') else '',
            f"${int(p['Participant_BLife_Amount']):,}" if p.get('Participant_BLife_Amount') not in (None, 0) else '',
            str(p.get('Participant_BLife_Dep_Units', '')).split('.')[0] if p.get('Participant_BLife_Dep_Units') not in (None, 0) else '',
            p.get('Participant_VLife_EffectiveDate').strftime('%m/%d/%Y') if p.get('Participant_VLife_EffectiveDate') else '',
            f"${int(p['Participant_VLife_EE_Amount']):,}" if p.get('Participant_VLife_EE_Amount') not in (None, 0) else '',
            f"${p['Participant_VLife_EE_Prem']:,.2f}" if p.get('Participant_VLife_EE_Prem') is not None else '',
            f"${int(p['Participant_VLife_Spouse_Amount']):,}" if p.get('Participant_VLife_Spouse_Amount') not in (None, 0) else '',
            f"${p['Participant_VLife_Spouse_Prem']:,.2f}" if p.get('Participant_VLife_Spouse_Prem') is not None else '',
            f"${int(p['Participant_VLife_Child_Amount']):,}" if p.get('Participant_VLife_Child_Amount') not in (None, 0) else '',
            f"${p['Participant_VLife_Child_Prem']:,.2f}" if p.get('Participant_VLife_Child_Prem') is not None else '',
            f"${p['Participant_VLife_Total']:,.2f}" if p.get('Participant_VLife_Total') is not None else '',
            f"${p['Participant_Weekly_VL_Deduction']:,.2f}" if p.get('Participant_Weekly_VL_Deduction') is not None else '',
            p.get('Participant_TermDate').strftime('%m/%d/%Y') if p.get('Participant_TermDate') else ''
        ]
        data.append(row)
    table = Table(data, repeatRows=2, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # default left alignment
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Right-align dollar amount columns
        ('ALIGN', (4, 2), (4, -1), 'RIGHT'),   # BASIC COVERAGE Amt
        ('ALIGN', (7, 2), (7, -1), 'RIGHT'),   # VLife EE Amount
        ('ALIGN', (8, 2), (8, -1), 'RIGHT'),   # VLife EE Prem
        ('ALIGN', (9, 2), (9, -1), 'RIGHT'),   # Spouse Amount
        ('ALIGN', (10, 2), (10, -1), 'RIGHT'), # Spouse Prem
        ('ALIGN', (11, 2), (11, -1), 'RIGHT'), # Child Amount
        ('ALIGN', (12, 2), (12, -1), 'RIGHT'), # Child Prem
        ('ALIGN', (13, 2), (13, -1), 'RIGHT'), # VLife Total
        ('ALIGN', (14, 2), (14, -1), 'RIGHT'), # Weekly Deduct

        ('RIGHTPADDING', (4, 2), (4, -1), 2),
        ('RIGHTPADDING', (7, 2), (7, -1), 2),
        ('RIGHTPADDING', (8, 2), (8, -1), 2),
        ('RIGHTPADDING', (9, 2), (9, -1), 2),
        ('RIGHTPADDING', (10, 2), (10, -1), 2),
        ('RIGHTPADDING', (11, 2), (11, -1), 2),
        ('RIGHTPADDING', (12, 2), (12, -1), 2),
        ('RIGHTPADDING', (13, 2), (13, -1), 2),
        ('RIGHTPADDING', (14, 2), (14, -1), 2),

        # Header styling
        ('SPAN', (3, 0), (5, 0)),
        ('SPAN', (6, 0), (8, 0)),
        ('SPAN', (9, 0), (10, 0)),
        ('SPAN', (11, 0), (12, 0)),
        ('BACKGROUND', (0, 0), (-1, 1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))

    story.append(table)

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)

    return Response(
        buffer.read(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "inline; filename=insurance_monthly.pdf"}
    )

def generate_payroll_deduction_analysis_110(participants, pmt_weeks, month_year, session):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=20 * mm
    )

    styles = getSampleStyleSheet()
    story = []
    timestamp = datetime.now().strftime("%B %d, %Y %I:%M %p")

    story.append(Paragraph("Payroll Deduction Analysis", styles["Title"]))
    story.append(Paragraph(f"Generated on: {timestamp}", styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    header1 = [
        'Name', 'ID',
        'BASIC COVERAGE', '',
        'VOLUNTARY LIFE', '', '', '', '', '', '',
        'PAYROLL DEDUCTIONS', '', '', '', ''
    ]
    header2 = [
        '', '',
        'Effective', 'Group',
        'Eff Date', 'End Date', 'Emp Prem', 'Spouse Prem',
        'Child Prem', 'Total', 'Weekly Deduct',
        'Month', 'WKS', 'Due', 'Paid', 'Total'
    ]

    data = [header1, header2]

    for p in participants:
        weekly_deduction = p.get('Participant_Weekly_VL_Deduction') or 0
        curr_date = datetime.strptime(month_year, "%m/%Y")
        due_date = add_one_month(curr_date).strftime("%Y-%m-%d")
        participant_pkey = p.get('Participant_Pkey')

        paid_amt = session.execute(text('''
            SELECT Prem_Payroll_Deduction
            FROM XParticipant_Billing
            WHERE Prem_Pkey = :pkey AND Prem_Due_Date = :due_date
        '''), {'pkey': participant_pkey, 'due_date': due_date}).scalar() or 0        
        due_amt = weekly_deduction * pmt_weeks
        total_amt = due_amt - paid_amt

        def fmt(val):
            return f"${val:,.2f}  " if val is not None else ''

        row = [
            f"{p.get('Participant_FirstName', '')} {p.get('Participant_LastName', '')}",
            str(p.get('Participant_SSN'))[-4:] if p.get('Participant_SSN') else '',
            p.get('Participant_BLife_EffectiveDate').strftime('%m/%d/%Y') if p.get('Participant_BLife_EffectiveDate') else '',
            p.get('Participant_Group_Name', ''),
            p.get('Participant_VLife_EffectiveDate').strftime('%m/%d/%Y') if p.get('Participant_VLife_EffectiveDate') else '',
            p.get('Participant_VLife_EndDate').strftime('%m/%d/%Y') if p.get('Participant_VLife_EndDate') else '',
            fmt(p.get('Participant_VLife_EE_Prem')),
            fmt(p.get('Participant_VLife_Spouse_Prem')),
            fmt(p.get('Participant_VLife_Child_Prem')),
            fmt(p.get('Participant_VLife_Total')),
            fmt(weekly_deduction),
            month_year,
            str(pmt_weeks),
            fmt(due_amt),
            fmt(paid_amt),
            fmt(total_amt),
        ]
        data.append(row)

    table = Table(data, repeatRows=2)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Header spans
        ('SPAN', (0, 0), (0, 1)),  # Name
        ('SPAN', (1, 0), (1, 1)),  # ID
        ('SPAN', (2, 0), (3, 0)),  # BASIC COVERAGE
        ('SPAN', (4, 0), (10, 0)), # VOLUNTARY LIFE
        ('SPAN', (11, 0), (15, 0)),# PAYROLL DEDUCTIONS

        # Header background and border
        ('BACKGROUND', (0, 0), (-1, 1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),

        # Padding
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (6, 2), (10, -1), 1),
        ('RIGHTPADDING', (13, 2), (15, -1), 1),

        # Right-align all $ columns
        ('ALIGN', (6, 2), (10, -1), 'RIGHT'),   # VLife
        ('ALIGN', (13, 2), (15, -1), 'RIGHT'),  # Due, Paid, Total
    ]))

    story.append(table)

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)

    return Response(
        buffer.read(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "inline; filename=payroll_deduction_analysis.pdf"}
    )


def generate_summary_120_pdf(participants, family_members):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer)

    full_story = []

    for p in participants:
        dependents = family_members.get(p['Participant_Pkey'], [])
        full_story.extend(build_summary(p, dependents))  # append story for this participant


    if full_story and isinstance(full_story[-1], PageBreak):
        full_story.pop()  # Remove trailing PageBreak if last item

    doc.build(full_story, onFirstPage=add_page_number, onLaterPages=add_page_number)

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={
            "Content-Disposition": "attachment; filename=participant_summary.pdf"
        }
    )


def build_summary(p, dependents):

    def fmt(dt):
        return dt.strftime('%m/%d/%Y') if dt else ''

    def ssn_format(ssn):
        try:
            ssn = str(int(float(ssn))) 
        except (ValueError, TypeError):
            return str(ssn) 
        return f"{ssn[:3]}-{ssn[3:5]}-{ssn[5:]}" if len(ssn) == 9 else ssn

    styles = getSampleStyleSheet()
    story = []

    name = f"{p['Participant_LastName']}, {p['Participant_FirstName']}"
    ssn = ssn_format(p.get('Participant_SSN'))
    dob = fmt(p.get('Participant_BirthDate'))
    id_num = p.get('Participant_ID')
    age_val = p.get('Participant_Age')
    age = str(age_val).split('.')[0] if age_val is not None else ''
    term = fmt(p.get('Participant_TermDate'))
    begin = fmt(p.get('Participant_BeginDate'))
    end = fmt(p.get('Participant_EndDate'))
    access_point = p.get('PPO_AccessPoint1') or ''
    network = p['PPO_Network_Display'] = (
        f"{p.get('Participant_Pkey')}"
        if not p.get('PPO_Network1') and not access_point
        else f"{p.get('PPO_Network1', '')} | Access Point: {access_point}"
    )

    med_start = fmt(p.get('Participant_MED_EffectiveDate'))
    med_end = fmt(p.get('Participant_MED_EndDate'))
    dental_start = fmt(p.get('Participant_DEN_EffectiveDate'))
    dental_end = fmt(p.get('Participant_DEN_EndDate'))
    vision_start = fmt(p.get('Participant_VIS_EffectiveDate'))
    vision_end = fmt(p.get('Participant_VIS_EndDate'))

    blife_start = fmt(p.get('Participant_BLife_EffectiveDate'))
    blife_end = fmt(p.get('Participant_BLife_EndDate'))
    vlife_start = fmt(p.get('Participant_VLife_EffectiveDate'))
    vlife_end = fmt(p.get('Participant_VLife_EndDate'))

    # Header
    story.append(Paragraph("<b>Full Participant Report</b>", styles["Title"]))
    story.append(Paragraph(f"<b>{name}</b>", styles["Heading2"]))
    story.append(Paragraph(f"ID# {id_num}", styles["Normal"]))
    story.append(Spacer(1, 12))

    timestamp = datetime.now().strftime("%B %d, %Y %I:%M %p")
    story.append(Paragraph(f"Generated on: {timestamp}", styles["Normal"]))
    # Left: Participant + Healthcare Info

    # Participant Information header
    tight_padding = [
        ("LEFTPADDING", (0, 0), (-1, -1), 0.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0.5),
        ("TOPPADDING", (0, 0), (-1, -1), 0.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0.5),
    ]

    # Participant Information header
    part_info_flowable = [
        Table(
            [[Paragraph("<b>Participant Information</b>", styles["Heading4"])]],
            colWidths=[3.0 * inch],
            style=TableStyle([
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                *tight_padding,
            ])
        )
    ]

    # Participant data table
    part_info_table_data = [
        ["Name", name],
        ["Address", p.get("Participant_Address", "")],
        ["City/State", f"{p.get('Participant_City', '')}, {p.get('Participant_State', '')} {p.get('Participant_Zip', '')}"],
        ["Date-Hired", fmt(p.get("Participant_HireDate")), "Termin.", term],
        ["Birthdate", dob, "Age", age],
        ["ID#", id_num, "Status", p.get("Participant_Status", "")]
    ]

    part_info_flowable.append(
        Table(part_info_table_data, colWidths=[0.9*inch, 2.1*inch, 0.9*inch, 0.9*inch], style=TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            *tight_padding,
        ]))
    )

    # Healthcare Information header
    healthcare_flowable = [
        Table(
            [[Paragraph("<b>Healthcare Information</b>", styles["Heading4"])]],
            colWidths=[3.0 * inch],
            style=TableStyle([
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                *tight_padding,
            ])
        )
    ]

    # Healthcare data table
    healthcare_table_data = [
        ["Group-Name", p.get("Participant_Group_Name", ""), "Division", p.get("Participant_Division", "")],
        ["Begin-Date", begin, "Ending", end],
        ["Special Packet", fmt(p.get("Participant_MailDate_Packet")), "", ""],
        ["Convers. Port", fmt(p.get("Participant_MailDate_Convert_Port")), "Life Class", p.get("Participant_LIFE_CLASS")],
        ["CERTs. Mailed", fmt(p.get("Participant_MailDate_CERTs")), "Life Only", p.get("Participant_LIFE_ONLY_YN", "N")],
        ["PPO Network: ", network, "", ""] 
    ]

    healthcare_flowable.append(
        Table(healthcare_table_data, colWidths=[0.9*inch, 1.1*inch, 0.9*inch, 0.9*inch], style=TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("SPAN", (1, 5), (3, 5)),
            *tight_padding,
        ]))
    )


    left_column_data = []

    # Participant Info Section
    left_column_data.append([Paragraph("<b>Participant Information</b>", styles["Heading4"]), '', '', ''])
    left_column_data.extend([
        ["Name", name, "Birthdate", dob],
        ["Address", p.get("Participant_Address", ""), "Age", age],
        ["City/State", f"{p.get('Participant_City', '')}, {p.get('Participant_State', '')} {p.get('Participant_Zip', '')}", "ID#", id_num],
        ["Date-Hired", fmt(p.get("Participant_HireDate")), "Termin.", term],
        ["Status", p.get("Participant_Status", ""), '', ''],
    ])

    left_column_data.append(["", "", "", ""]) 
    spacer_index = len(left_column_data) - 1
    # Healthcare Info Section
    # left_column_data.append(['', '', '', ''])  
    left_column_data.append([Paragraph("<b>Healthcare Information</b>", styles["Heading4"]), '', '', ''])
    left_column_data.extend([
        ["Group-Name", p.get("Participant_Group_Name", ""), "Division", p.get("Participant_Division", "")],
        ["Begin-Date", begin, "Ending", end],
        ["Special Packet", "", "", ""],
        ["Convers. Port", fmt(p.get("Participant_MailDate_Convert_Port")), "Life Class", p.get("Participant_LIFE_CLASS")],
        ["CERTs. Mailed", fmt(p.get("Participant_MailDate_CERTs")), "Life Only", p.get("Participant_LIFE_ONLY_YN", "N")],
        ["PPO Network", network, "", ""],
    ])

    left_col_table = Table(
        left_column_data,
        colWidths=[1.2 * inch, 1.8 * inch, 1.0 * inch, 1.0 * inch],
        style=TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),  # "Participant Information"
            ("SPAN", (0, 0), (3, 0)),  # Span the top header row
            ("SPAN", (0, 7), (3, 7)),  # "Healthcare Information" (was at 6)
            ("BACKGROUND", (0, 7), (3, 7), colors.lightgrey),
            ("LINEBEFORE", (0, spacer_index), (-1, spacer_index), 0, colors.white),
            ("LINEAFTER", (0, spacer_index), (-1, spacer_index), 0, colors.white),
            ("BACKGROUND", (0, spacer_index), (-1, spacer_index), colors.white),
        ])

    )
    #coverage section
    coverage_table_data = [
        [Paragraph("<b>Coverage Information</b>", styles["Heading4"]), "", ""],  # Header row, spans all columns
        ["", "Begin-Date", "Ending Date"],
        ["HEALTHCARE", "", ""],
        ["Medical", med_start, med_end],
        ["Vision", vision_start, vision_end],
        ["Dental", dental_start, dental_end],
        ["BASIC LIFE", "", ""],
        ["Employee", blife_start, blife_end],
        [
            "Spouse",
            p['Participant_BLife_EffectiveDate'].strftime("%m/%d/%Y") if (p.get('Participant_BLife_Spouse_Amount') or 0) > 0 and p.get('Participant_BLife_EffectiveDate') else "",
            p['Participant_BLife_EndDate'].strftime("%m/%d/%Y") if (p.get('Participant_BLife_Spouse_Amount') or 0) > 0 and p.get('Participant_BLife_EndDate') else ""
        ],
        [
            "Children",
            p['Participant_BLife_EffectiveDate'].strftime("%m/%d/%Y") if (p.get('Participant_BLife_Child_Amount') or 0) > 0 and p.get('Participant_BLife_EffectiveDate') else "",
            p['Participant_BLife_EndDate'].strftime("%m/%d/%Y") if (p.get('Participant_BLife_Child_Amount') or 0) > 0 and p.get('Participant_BLife_EndDate') else ""
        ],
        ["VOL LIFE", "", ""],
        ["Employee", vlife_start, vlife_end],
        [
            "Spouse", 
            p['Participant_VLife_EffectiveDate'].strftime("%m/%d/%Y") if (p.get('Participant_VLife_Spouse_Amount') or 0) > 0 and p.get('Participant_VLife_EffectiveDate') else "",
            p['Participant_VLife_EndDate'].strftime("%m/%d/%Y") if (p.get('Participant_VLife_Spouse_Amount') or 0) > 0 and p.get('Participant_VLife_EndDate') else ""
        ],
        [
            "Children", 
            p['Participant_VLife_EffectiveDate'].strftime("%m/%d/%Y") if (p.get('Participant_VLife_Child_Amount') or 0) > 0 and p.get('Participant_VLife_EffectiveDate') else "",
            p['Participant_VLife_EndDate'].strftime("%m/%d/%Y") if (p.get('Participant_VLife_Child_Amount') or 0) > 0 and p.get('Participant_VLife_EndDate') else ""
        ]
    ]

    right_col_table = Table(
        coverage_table_data,
        colWidths=[1.3 * inch, 0.85 * inch, 0.85 * inch],
        style=TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

            # Header row (Coverage Information)
            ("BACKGROUND", (0, 0), (2, 0), colors.lightgrey),
            ("ALIGN", (0, 0), (2, 0), "CENTER"),
            ("SPAN", (0, 0), (2, 0)),

            # Subheader row
            ("FONTNAME", (0, 1), (2, 1), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (2, 1), "CENTER"),

            # Section titles
            ("SPAN", (0, 2), (2, 2)),  # HEALTHCARE
            ("SPAN", (0, 6), (2, 6)),  # BASIC LIFE
            ("SPAN", (0, 10), (2, 10)),  # VOL LIFE

            ("FONTNAME", (0, 2), (0, 2), "Helvetica-Bold"),
            ("FONTNAME", (0, 6), (0, 6), "Helvetica-Bold"),
            ("FONTNAME", (0, 10), (0, 10), "Helvetica-Bold"),

            ("ALIGN", (0, 2), (2, 2), "LEFT"),
            ("ALIGN", (0, 6), (2, 6), "LEFT"),
            ("ALIGN", (0, 10), (2, 10), "LEFT"),
        ])
    )


    # Place the two stacked tables side-by-side in one large row-wise table
    story.append(
        Table(
            [[left_col_table, Spacer(1, 1), right_col_table]],
            colWidths=[5 * inch, 0.1 * inch, 3 * inch],
            style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ])
        )
    )

    story.append(Spacer(1, 8))
    # Family
    # Family Section Header Rows

    life_only_label = "Life Only" if p.get("Participant_LIFE_ONLY_YN") == "Y" else "Medical"

    family_info = [
        ["Family Profile", "", "Coverage Effective Dates", "", "", "", ""], 
        ["Name", "Relationship", life_only_label, "Dental", "Vision", "End Date", "PPO Network"] 
    ]

    # Row spans
    span_commands = [
        ('SPAN', (0, 0), (1, 0)),  # Family Profile spans columns 0–1
        ('SPAN', (2, 0), (6, 0)),  # Coverage Effective Dates spans columns 2–6
        ('BACKGROUND', (0, 0), (6, 0), colors.lightgrey),
        ('BACKGROUND', (0, 1), (6, 1), colors.HexColor("#f0f0f0")),
        ('FONTNAME', (0, 0), (6, 1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (6, 1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
    ]

    # Participant (employee) row

    family_info.append([
        f"{p['Participant_FirstName']} {p['Participant_LastName']}",
        "Employee",
        fmt(
            p.get("Participant_BeginDate") if (
                p.get("Participant_LIFE_ONLY_YN") == "Y" and not p.get("Participant_VLife_EffectiveDate")
            ) else
            p.get("Participant_VLife_EffectiveDate") if (
                p.get("Participant_LIFE_ONLY_YN") == "Y"
            ) else
            p.get("Participant_MED_EffectiveDate")
        ),
        fmt(p.get("Participant_DEN_EffectiveDate")),
        fmt(p.get("Participant_VIS_EffectiveDate")),
        fmt(
            p.get("Participant_VLife_EndDate") if p.get("Participant_LIFE_ONLY_YN") == "Y"
            else p.get("Participant_MED_EndDate")
        ),
        p.get("Participant_EE_PPO"),
    ])

    # Dependents
    for d in dependents:
        dep_name = f"{d['Depend_FirstName']} {d['Depend_LastName']}"
        rel = d.get("Depend_Relation", "")
        if p.get("Participant_LIFE_ONLY_YN") == "Y":
            med = fmt(p.get("Participant_VLife_EffectiveDate"))
        elif d.get("Depend_InEligible_YN"):
            med = "NE"
        else:
            med = fmt(d.get("Depend_MED_Effective"))
        den = fmt(d.get("Depend_DEN_Effective"))
        vis = fmt(d.get("Depend_VIS_Effective"))
        end_date = "" if d.get("Depend_InEligible_YN") else fmt(d.get("Depend_MED_EndDate"))
        ppo = d.get("Depend_PPO", "")
        family_info.append([dep_name, rel, med, den, vis, end_date, ppo])

    story.append(Paragraph("<b>Family Profile | Coverage Effective Dates</b>", styles["Heading4"]))
    story.append(Table(family_info, colWidths=[
        1.8*inch, 1.1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1.2*inch,
    ], style=TableStyle(span_commands)))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Notes", styles["Heading4"]))
    notes = p.get("Participant_Coverage_Notes") or ""
    story.append(Paragraph(notes if notes.strip() else "", styles["Normal"]))
    story.append(PageBreak())

    return story


def generate_age_reduction_115(results):
    from collections import defaultdict
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=20 * mm
    )

    styles = getSampleStyleSheet()
    story = []
    timestamp = datetime.now().strftime("%B %d, %Y %I:%M %p")

    story.append(Paragraph("Age Reduction", styles["Title"]))
    story.append(Paragraph(f"Generated on: {timestamp}", styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    header1 = [
        'Age', 'Birthday', 'Name', 'ID', 'Group',
        'BASIC LIFE - Employee', '',
        'VOLUNTARY LIFE - Employee', '', '',
        'Spouse', '',
        'Child', '',
        'VLife Total', 'Weekly Deduct'
    ]
    header2 = [
        '', '', '', '', '',
        'Effective', 'Amt',
        'Effective', 'Amount', 'Prem',
        'Amount', 'Prem',
        'Amount', 'Prem',
        'Total', 'Deduct'
    ]
    data = [header1, header2]
    span_commands = []

    # Add header spans
    span_commands += [
        ('SPAN', (0, 0), (0, 1)),  # Age
        ('SPAN', (1, 0), (1, 1)),  # Birthday
        ('SPAN', (2, 0), (2, 1)),  # Name
        ('SPAN', (3, 0), (3, 1)),  # ID
        ('SPAN', (4, 0), (4, 1)),  # Group
        ('SPAN', (5, 0), (6, 0)),  # Basic Life
        ('SPAN', (7, 0), (9, 0)),  # VLife EE
        ('SPAN', (10, 0), (11, 0)),  # Spouse
        ('SPAN', (12, 0), (13, 0)),  # Child
        ('SPAN', (14, 0), (14, 1)),  # VLife Total
        ('SPAN', (15, 0), (15, 1)),  # Weekly Deduct
    ]

    # Group and sort results
    results = sorted(results, key=lambda r: r['Participant_Age'] or 0)
    age_counts = defaultdict(int)
    for r in results:
        age_counts[r['Participant_Age']] += 1

    last_age = None
    row_idx = 2  # Starting index of data (after 2 header rows)

    for p in results:
        current_age = p['Participant_Age']

        if current_age != last_age:
            count = age_counts[current_age]
            label = f"({count}) Participant{'s' if count != 1 else ''} Currently Age {current_age if current_age is not None else 'Unknown'} Qualify For Age Reduction"
            divider_row = [''] * 16
            divider_row[0] = label
            data.append(divider_row)

            span_commands += [
                ('SPAN', (0, row_idx), (15, row_idx)),
                ('BACKGROUND', (0, row_idx), (15, row_idx), colors.HexColor("#dee2e6")),
                ('ALIGN', (0, row_idx), (15, row_idx), 'CENTER'),
                ('FONTNAME', (0, row_idx), (15, row_idx), 'Helvetica-Bold'),
            ]
            row_idx += 1
            last_age = current_age

        row = [
            str(p['Participant_Age']).split('.')[0] if p['Participant_Age'] is not None else '',
            p['Participant_BirthDate'].strftime('%m/%d/%Y') if p['Participant_BirthDate'] else '',
            f"{p['Participant_FirstName']} {p['Participant_LastName']}",
            str(p['Participant_SSN'])[-4:] if p['Participant_SSN'] else '',
            p.get('Participant_Group_Name', ''),
            p['Participant_BLife_EffectiveDate'].strftime('%m/%d/%Y') if p.get('Participant_BLife_EffectiveDate') else '',
            f"${int(p['Participant_BLife_Amount']):,}" if p.get('Participant_BLife_Amount') not in (None, 0) else '',
            p['Participant_VLife_EffectiveDate'].strftime('%m/%d/%Y') if p.get('Participant_VLife_EffectiveDate') else '',
            f"${int(p['Participant_VLife_EE_Amount']):,}" if p.get('Participant_VLife_EE_Amount') not in (None, 0) else '',
            f"${p['Participant_VLife_EE_Prem']:,.2f}" if p.get('Participant_VLife_EE_Prem') is not None else '',
            f"${int(p['Participant_VLife_Spouse_Amount']):,}" if p.get('Participant_VLife_Spouse_Amount') not in (None, 0) else '',
            f"${p['Participant_VLife_Spouse_Prem']:,.2f}" if p.get('Participant_VLife_Spouse_Prem') is not None else '',
            f"${int(p['Participant_VLife_Child_Amount']):,}" if p.get('Participant_VLife_Child_Amount') not in (None, 0) else '',
            f"${p['Participant_VLife_Child_Prem']:,.2f}" if p.get('Participant_VLife_Child_Prem') is not None else '',
            f"${p['Participant_VLife_Total']:,.2f}" if p.get('Participant_VLife_Total') is not None else '',
            f"${p['Participant_Weekly_VL_Deduction']:,.2f}" if p.get('Participant_Weekly_VL_Deduction') is not None else ''
        ]
        data.append(row)
        row_idx += 1

    table = Table(data, repeatRows=2)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Right-align dollar amounts
        ('ALIGN', (6, 2), (6, -1), 'RIGHT'),    # BASIC Amt
        ('ALIGN', (8, 2), (8, -1), 'RIGHT'),    # VLife EE Amt
        ('ALIGN', (9, 2), (9, -1), 'RIGHT'),    # VLife EE Prem
        ('ALIGN', (10, 2), (10, -1), 'RIGHT'),  # Spouse Amt
        ('ALIGN', (11, 2), (11, -1), 'RIGHT'),  # Spouse Prem
        ('ALIGN', (12, 2), (12, -1), 'RIGHT'),  # Child Amt
        ('ALIGN', (13, 2), (13, -1), 'RIGHT'),  # Child Prem
        ('ALIGN', (14, 2), (14, -1), 'RIGHT'),  # Total
        ('ALIGN', (15, 2), (15, -1), 'RIGHT'),  # Deduct

        # Tighten spacing to hug right edge
        ('RIGHTPADDING', (6, 2), (6, -1), 2),
        ('RIGHTPADDING', (8, 2), (8, -1), 2),
        ('RIGHTPADDING', (9, 2), (9, -1), 2),
        ('RIGHTPADDING', (10, 2), (10, -1), 2),
        ('RIGHTPADDING', (11, 2), (11, -1), 2),
        ('RIGHTPADDING', (12, 2), (12, -1), 2),
        ('RIGHTPADDING', (13, 2), (13, -1), 2),
        ('RIGHTPADDING', (14, 2), (14, -1), 2),
        ('RIGHTPADDING', (15, 2), (15, -1), 2),

        ('BACKGROUND', (0, 0), (-1, 1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        *span_commands
    ]))

    story.append(table)

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)

    return Response(
        buffer.read(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "inline; filename=age_reduction.pdf"}
    )

def build_employer_report(p, dependents):
    styles = getSampleStyleSheet()
    story = []

    def fmt(dt):
        return dt.strftime('%m/%d/%Y') if dt else ''

    name = f"{p['Participant_FirstName']} {p['Participant_LastName']}"
    dept = p.get("Participant_Department", "")
    id_num = p.get("Participant_ID", "")
    hire_date = fmt(p.get("Participant_HireDate"))
    birth_date = fmt(p.get("Participant_BirthDate"))
    med_eff = fmt(p.get("Participant_MED_EffectiveDate"))
    dental_eff = fmt(p.get("Coverage_Dental_BeginDate"))
    vision_eff = fmt(p.get("Coverage_Vision_BeginDate"))
    term_date = fmt(p.get("Participant_TermDate"))
    address = p.get("Participant_Address", "")
    city_state = f"{p.get('Participant_City', '')}, {p.get('Participant_State', '')} {p.get('Participant_Zip', '')}"
    division = p.get("Participant_Division", "")
    ppo = p.get("Participant_EE_PPO", "")
    comments = p.get("Participant_Coverage_Notes", "") or ""

    # --- Header bar: EMPLOYER REPORT | DEPARTMENT
    story.append(Paragraph(
        f"<b>EMPLOYER REPORT</b> &nbsp;&nbsp;&nbsp;&nbsp; | &nbsp;&nbsp;&nbsp;&nbsp; {dept}", styles["Heading2"]
    ))
    story.append(Spacer(1, 8))

    # --- ACTIVE row
    story.append(
        Table(
            [["Active", "Dept:", dept, "", "EMPLOYEE COVERAGE INFORMATION", "ID#", id_num]],
            colWidths=[0.7*inch, 0.7*inch, 1.1*inch, 0.5*inch, 3*inch, 0.5*inch, 1.2*inch],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (6, 0), colors.black),
                ("TEXTCOLOR", (0, 0), (6, 0), colors.white),
                ("FONTNAME", (0, 0), (6, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (6, 0), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ])
        )
    )

    story.append(Spacer(1, 2))

     # --- Employee coverage info table
    employee_info_data = [
        ["Name", name, "Medical-Effective", med_eff, "Hired", hire_date],
        ["Address", address, "Dental-Effective", dental_eff, "BirthDate", birth_date],
        ["City/State", city_state, "Vision-Effective", vision_eff, "PPO", ppo],
        ["Division", division, "", "", "Termed", term_date],
        ["Comments", Paragraph(comments, styles["Normal"]), "", "", "", ""],
    ]

    story.append(
        Table(
            employee_info_data,
            colWidths=[
                1.1 * inch, 2.1 * inch,
                1.2 * inch, 1.0 * inch,
                0.9 * inch, 1.0 * inch
            ],
            style=TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("BACKGROUND", (2, 0), (2, -1), colors.whitesmoke),
                ("BACKGROUND", (4, 0), (4, -1), colors.whitesmoke),
                ("SPAN", (1, 4), (5, 4)),  # Comments spans 3 cells
            ])
        )
    )

    story.append(Spacer(1, 6))

    # --- Checkbox Grid (labels on top, checkboxes underneath)
    def checked(field):
        val = p.get(field)
        return "✓" if str(val).strip().upper() in {"Y", "1", "TRUE"} else ""

    checkbox_labels = [
        "Basic Life", "Opt Out", "Med / RX", "Dental", "Vision",
        "Cobra", "Waived VLife", "Waived Hth&Life", "Guar. Pay"
    ]

    checkbox_values = [
        checked("Participant_Coverage_BasicLife_YN"),
        checked("Participant_Coverage_OptOut_YN"),
        checked("Participant_Coverage_MedRX_YN"),
        checked("Participant_Coverage_Dental_YN"),
        checked("Participant_Coverage_Vision_YN"),
        checked("Participant_Coverage_Cobra_YN"),
        checked("Participant_VLife_Waived_YN"),
        checked("Participant_Coverage_Waived_YN"),
        checked("Participant_Coverage_GuaranteePay_YN")
    ]

    checkbox_data = [checkbox_labels, checkbox_values]

    story.append(
        Table(
            checkbox_data,
            colWidths = [
                0.78 * inch,  # Basic Life
                0.78 * inch,  # Opt Out
                0.78 * inch,  # Med / RX
                0.78 * inch,  # Dental
                0.78 * inch,  # Vision
                0.78 * inch,  # Cobra
                0.85 * inch,  # Waived VLife
                1.1 * inch,   # Waived Hth&Life 
                0.87 * inch   # Guar. Pay
            ],
            style=TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
                ("TOPPADDING", (0, 1), (-1, 1), 1),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
            ])
        )
    )

    story.append(Spacer(1, 6))

        # --- Family & Dependents Table
    if dependents:
        story.append(
            Table(
                [[
                    Paragraph("<b>Family Members & Dependents</b>", styles["Normal"]),
                    "", "", "", "", "", "", Paragraph(f"<b>({len(dependents)})</b>", styles["Normal"])
                ]],
                colWidths=[1.5*inch, 0.6*inch, 0.8*inch, .8*inch, .9*inch, .9*inch, 0.7*inch, 1.5*inch],
                style=TableStyle([
                    ("SPAN", (0, 0), (5, 0)),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 4)
                ])
            )
        )

        dep_headers = [
            "Name", "Relation", "Birthdate", "Med-Effective",
            "Dental-Effective", "Vision-Effective", "Termed", "PPO-Network"
        ]
        story.append(
            Table(
                [dep_headers] + [[
                    d.get("Depend_Name", ""),
                    d.get("Depend_Relation", ""),
                    fmt(d.get("Depend_BirthDate")),
                    fmt(d.get("Depend_MED_Effective")), 
                    fmt(d.get("Depend_DEN_Effective")),    
                    fmt(d.get("Depend_VIS_Effective")),
                    fmt(d.get("Depend_TermDate")),
                    d.get("Depend_Pkey", "")
                ] for d in dependents],
                colWidths=[1.5*inch, 0.6*inch, 0.8*inch, .8*inch, .9*inch, .9*inch, 0.7*inch, 1.5*inch],
                style=TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ])
            )
        )

        story.append(Spacer(1, 10))

    return story


def build_new_employer_report_header():
    header_style = ParagraphStyle("HeaderStyle", fontName="Helvetica-Bold", fontSize=16, alignment=TA_CENTER, spaceAfter=4)
    subheader_style = ParagraphStyle("SubHeaderStyle", fontName="Helvetica-Bold", fontSize=12, alignment=TA_CENTER, spaceAfter=8)
    paragraph_style = ParagraphStyle("BodyStyle", fontName="Helvetica", fontSize=9, alignment=0, leading=12, spaceAfter=6)

    return [
        Paragraph("PARTICIPANT REPORT", header_style),
        HRFlowable(width="100%", thickness=1, color=colors.black, spaceBefore=4, spaceAfter=8),
        Paragraph("REPORT OF COVERAGE STATUS", subheader_style),
        Paragraph('This “Report of Coverage” is not a guarantee of coverage or benefits and is for informational purpose only.', paragraph_style),
        Paragraph('Please check this report for accuracy and let us know if any changes are needed.', paragraph_style),
        Paragraph('Please contact the insurance clerk at your place of employment or Eric Kiser Insurance if you have any questions about the coverage status of a specific person(s) or about this report.', paragraph_style),
        Spacer(1, 12),
    ]

def build_employee_address_section(participant, styles):
    left_data = [
        ["<b>Name</b>", f'{participant.get("Participant_LastName", "")}, {participant.get("Participant_FirstName", "")}'],
        ["<b>Address</b>", participant.get("Participant_Address", "")],
        ["<b>City/State</b>", f'{participant.get("Participant_City", "")}, {participant.get("Participant_State", "")}'],
    ]

    right_data = [
        ["<b>Company</b>", f"{participant.get('Participant_Group_Name', '')} - Dept - {participant.get('Participant_Department', '')}"],
        ["<b>Division</b>", participant.get("Participant_Division", "")],
    ]

    left_table = Table(
        [[Paragraph(cell if cell is not None else '', styles["Normal"]) for cell in row] for row in left_data],
        colWidths=[0.9 * inch, 2.3 * inch],
        style=TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )

    right_table = Table(
        [[Paragraph(str(cell) if cell is not None else '', styles["Normal"]) for cell in row] for row in right_data],
        colWidths=[0.9 * inch, 2.2 * inch],
        style=TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )

    # Combine both tables horizontally
    combined_table = Table(
        [[left_table, right_table]],
        colWidths=[3.3 * inch, 3.1 * inch]
    )

    title = Paragraph("<b><u>Employee Name and Address</u></b>", styles["Normal"])

    return [Spacer(1, 6), title, Spacer(1, 4), combined_table, Spacer(1, 12)]

def build_family_profile_table(p, dependents):

    def fmt(dt):
        return dt.strftime('%m/%d/%Y') if isinstance(dt, datetime) else (dt if dt else "")

    life_only_label = "Life Only" if p.get("Participant_LIFE_ONLY_YN") == "Y" else "Medical"

    family_info = [
        ["Family Profile", "", "Coverage Effective Dates", "", "", "", ""], 
        ["Name", "Relationship", life_only_label, "Dental", "Vision", "End Date", "PPO Network"] 
    ]

    span_commands = [
        ('SPAN', (0, 0), (1, 0)),
        ('SPAN', (2, 0), (6, 0)),
        ('BACKGROUND', (0, 0), (6, 0), colors.lightgrey),
        ('BACKGROUND', (0, 1), (6, 1), colors.HexColor("#f0f0f0")),
        ('FONTNAME', (0, 0), (6, 1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (6, 1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
    ]

    family_info.append([
        f"{p.get('Participant_FirstName', '')} {p.get('Participant_LastName', '')}",
        "Employee",
        fmt(
            p.get("Participant_BeginDate") if (
                p.get("Participant_LIFE_ONLY_YN") == "Y" and not p.get("Participant_VLife_EffectiveDate")
            ) else
            p.get("Participant_VLife_EffectiveDate") if (
                p.get("Participant_LIFE_ONLY_YN") == "Y"
            ) else
            p.get("Participant_MED_EffectiveDate")
        ),
        fmt(p.get("Participant_DEN_EffectiveDate")),
        fmt(p.get("Participant_VIS_EffectiveDate")),
        fmt(
            p.get("Participant_VLife_EndDate") if p.get("Participant_LIFE_ONLY_YN") == "Y"
            else p.get("Participant_MED_EndDate")
        ),
        p.get("Participant_EE_PPO", "")
    ])

    for d in dependents:
        dep_name = f"{d.get('Depend_FirstName', '')} {d.get('Depend_LastName', '')}"
        rel = d.get("Depend_Relation", "")
        if p.get("Participant_LIFE_ONLY_YN") == "Y":
            med = fmt(p.get("Participant_VLife_EffectiveDate"))
        elif d.get("Depend_InEligible_YN"):
            med = "NE"
        else:
            med = fmt(d.get("Depend_MED_Effective"))
        den = fmt(d.get("Depend_DEN_Effective"))
        vis = fmt(d.get("Depend_VIS_Effective"))
        end_date = "" if d.get("Depend_InEligible_YN") else fmt(d.get("Depend_MED_EndDate"))
        ppo = d.get("Depend_PPO", "")
        family_info.append([dep_name, rel, med, den, vis, end_date, ppo])

    return [
        Paragraph("<b><u>Family Profile | Coverage Effective Dates</u></b>"),
        Spacer(1, 4),
        Table(family_info, colWidths=[
            1.5*inch, 1*inch, 0.9*inch, .9*inch, .9*inch, .9*inch, 1.1*inch,
        ], style=TableStyle(span_commands)),
        Spacer(1, 12)
    ]

def generate_new_employee_130(participants, family_members):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    story = []
    styles = getSampleStyleSheet()

    for p in participants:
        story.extend(build_new_employer_report_header())
        story.extend(build_employee_address_section(p, styles))
        dependents = family_members.get(p["Participant_Pkey"], [])
        story.extend(build_family_profile_table(p, dependents))

        footer_header = ParagraphStyle("FooterHeader", fontName="Helvetica-Bold", fontSize=9, spaceAfter=6, alignment=TA_LEFT)
        footer_body = ParagraphStyle("FooterBody", fontName="Helvetica", fontSize=8, leading=11, spaceAfter=6)
        footer_note = ParagraphStyle("FooterNote", fontName="Helvetica", fontSize=8, leading=10, spaceAfter=4)

        story.extend([
            Spacer(1, 12),
            Paragraph("<b>IMPORTANT:</b>", footer_header),
            Paragraph("Each person to be covered must satisfy and maintain the Plan’s eligibility requirements in order to be covered and to remain covered under the group health plan. This report is not a guarantee of coverage.", footer_body),

            Paragraph("<b>NE = Not Eligible for Coverage:</b>", footer_header),
            Paragraph("Information/documents (such as a date of birth, date of marriage, divorce decree, birth certificate) for this person were requested and (i) the employee did not respond to the request, or (ii) the employee did not submit the requested information/documents, or (iii) the information/document(s) submitted were not satisfactory.", footer_body),
            Paragraph("For details regarding not eligible individuals please contact our office.", footer_body),

            Paragraph("<b>Swift MD (Telemedicine)</b>", footer_header),
            Paragraph("All covered Plan members are eligible to use the SwiftMD telemedicine program. Detailed information on this benefit program is included in your packet. Please take advantage of this program!", footer_body),
            Paragraph("Your SwiftMD membership is effective: 4/1/2023", footer_note),
            
            Spacer(1, 6),
            Paragraph("Eric Kiser Insurance Services LLC", footer_body),
            Paragraph("2321 Kochs Lane STE B, Quincy, IL 62305-9828", footer_note),
            Paragraph("Phone: (217) 228-0856 or (800) 772-0034   |   Email: eric@erickiser.com", footer_note),
        ])
        story.append(PageBreak())

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=new_employer_report.pdf"}
    )

def add_one_month(dt):
    year = dt.year + (dt.month // 12)
    month = dt.month % 12 + 1
    day = min(dt.day, [31,
        29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28,
        31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return datetime(year, month, day)


def add_page_number(canvas, doc):
    page_num = canvas.getPageNumber()
    text = f"Page {page_num}"
    canvas.setFont("Helvetica", 9)
    canvas.drawCentredString(doc.pagesize[0] / 2.0, 0.5 * inch, text)  


def generate_mailing_03(participants):
    def fmt_date(dt):
        if not dt:
            return ''
        try:
            return datetime.strptime(str(dt), "%Y-%m-%d").strftime('%m/%d/%Y')
        except ValueError:
            try:
                return datetime.strptime(str(dt), "%Y-%m-%d %H:%M:%S").strftime('%m/%d/%Y')
            except ValueError:
                return str(dt)

    LIGHT_YELLOW = colors.HexColor("#FFFDE7")
    LIGHT_GREEN  = colors.HexColor("#E8F5E9")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.75 * inch
    )

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    normal_style = styles['Normal']

    story = []
    story.append(Paragraph("Life Conversion-Port Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 0.2 * inch))

    headers = [
        "", "Group", "Participant", "Employment Status", "Life Status",
        "Begin Date", "End Date", "Basic Begin", "Basic End",
        "VL Begin", "VL End", "Conv-Port Mailed"
    ]
    table_data = [headers]

    for idx, p in enumerate(participants, start=1):
        life_flag = str(p.get("Participant_Coverage_BasicLife_YN", "")).strip().upper()
        life_status = "ACTIVE" if life_flag in ("Y", "TRUE", "1", "ACTIVE") else "TERM" if life_flag in ("N", "FALSE", "0", "TERM") else ""

        row = [
            str(idx),
            p.get("Participant_Group_Name", ""),
            p.get("Name", ""),
            p.get("Participant_Status", ""),
            life_status,
            fmt_date(p.get("Participant_BeginDate")),
            fmt_date(p.get("Participant_EndDate")),
            fmt_date(p.get("Participant_BLife_EffectiveDate")),
            fmt_date(p.get("Participant_BLife_EndDate")),
            fmt_date(p.get("Participant_VLife_EffectiveDate")),
            fmt_date(p.get("Participant_VLife_EndDate")),
            fmt_date(p.get("Participant_MailDate_Convert_Port"))
        ]
        table_data.append(row)

    col_widths = [
        20,   # 
        74,   # Group
        100,  # Participant
        80,   # Employment Status
        50,   # Life Status
        58,   # Begin Date
        58,   # End Date
        58,   # Basic Begin
        52,   # Basic End
        58,   # VL Begin
        48,   # VL End
        70    # Conv-Port Mailed
    ]

    table = Table(table_data, repeatRows=1, colWidths=col_widths)

    style = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),

        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ])

    for r in range(1, len(table_data)):
        style.add('BACKGROUND', (5, r), (5, r), LIGHT_YELLOW)
        style.add('BACKGROUND', (6, r), (6, r), LIGHT_YELLOW)
        style.add('BACKGROUND', (11, r), (11, r), LIGHT_GREEN)

        if table_data[r][8]:
            style.add('TEXTCOLOR', (8, r), (8, r), colors.red)
        if table_data[r][10]:
            style.add('TEXTCOLOR', (10, r), (10, r), colors.red)

    table.setStyle(style)
    story.append(table)

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)

    buffer.seek(0)
    return Response(
        buffer.read(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "inline; filename=life_conversion_port.pdf"}
    )




def generate_mailing_05(participants):
    def fmt_date(dt):
        if not dt:
            return ''
        try:
            return datetime.strptime(str(dt), "%Y-%m-%d").strftime('%m/%d/%Y')
        except ValueError:
            try:
                return datetime.strptime(str(dt), "%Y-%m-%d %H:%M:%S").strftime('%m/%d/%Y')
            except ValueError:
                return str(dt)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.75 * inch
    )

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    normal_style = styles['Normal']

    story = []
    story.append(Paragraph("Dental Coverage Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 0.2 * inch))

    headers = [
        "", "Group", "Participant", "Begin Date", "Med Eff", "Dental Eff", "Term Date", "Status", "Med Notes", "Participant Notes"
    ]
    table_data = [headers]

    wrap_style = ParagraphStyle(
        name='WrapStyle',
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        alignment=1  
    )


    for idx, p in enumerate(participants, start=1):

        row = [
            str(idx),
            p.get("Participant_Group_Name", ""),
            p.get("Name", ""),
            fmt_date(p.get("Participant_BeginDate")),
            fmt_date(p.get("Participant_MED_EffectiveDate")),
            fmt_date(p.get("Participant_DEN_EffectiveDate")),
            fmt_date(p.get("Participant_TermDate")),
            p.get("Participant_Status", ""),
            Paragraph(p.get("Participant_Coverage_Notes") or "", wrap_style),
            Paragraph(p.get("Participant_Insurance_Notes") or "", wrap_style)
        ]
        table_data.append(row)

    col_widths = [
        20,   # Index
        74,   # Group
        100,  # Participant
        80,   # Begin Date
        50,   # Med Eff
        50,   # Dental Eff
        80,   # Term Date
        50,   # Status
        100,  # Med Notes
        100   # Participant Notes
    ]

    table = Table(table_data, repeatRows=1, colWidths=col_widths)

    style = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),

        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ])

    table.setStyle(style)
    story.append(table)

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)

    buffer.seek(0)
    return Response(
        buffer.read(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "inline; filename=dental_coverage.pdf"}
    )


def generate_mailing_06(participants):
    def fmt_date(dt):
        if not dt:
            return ''
        try:
            return datetime.strptime(str(dt), "%Y-%m-%d").strftime('%m/%d/%Y')
        except ValueError:
            try:
                return datetime.strptime(str(dt), "%Y-%m-%d %H:%M:%S").strftime('%m/%d/%Y')
            except ValueError:
                return str(dt)

    LIGHT_BLUE   = colors.HexColor("#E3F2FD")
    LIGHT_ORANGE = colors.HexColor("#FFF3E0")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.75 * inch
    )

    styles = getSampleStyleSheet()
    small_header = ParagraphStyle(
        "SmallHeader",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=13,
        alignment=1  # center
    )
    normal_style = styles['Normal']

    story = []
    story.append(Paragraph("SPD+SBC+Notices+Participant Report + Life Cover Letter + Life Certs + Conversion Portability forms + Swift MD + Rx Cards", small_header))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 0.18 * inch))

    headers = [
        "", "Group", "DIV", "Participant", "CL", "LO", "Status",
        "Swift Mailed", "Begin Date", "Notice Mailed", "Med Eff", "Dental Eff",
        "Rx-Card Mailed", "SPD Mailed", "Basic Eff", "VL Eff", "L Cts Mailed"
    ]
    table_data = [headers]

    for idx, p in enumerate(participants, start=1):
        life_flag = (p.get("Participant_LIFE_ONLY_YN") or "").strip().upper()
        life_only_val = "" if life_flag in ("N", "0", "FALSE") else (p.get("Participant_LIFE_ONLY_YN") or "")
        status_val = (p.get("Participant_Status") or "").strip().upper()
        status = "A" if status_val == "ACTIVE" else ("T" if status_val == "TERM" else "")

        row = [
            str(idx),
            p.get("Participant_Group_Name", ""),
            p.get("Participant_Division", ""),
            p.get("Name", ""),
            p.get("CL", ""), #Unknown
            life_only_val,
            status,
            fmt_date(p.get("Participant_SwiftMD_MailDate")),
            fmt_date(p.get("Participant_BeginDate")),
            fmt_date(p.get("Notice_Mailed")), #unknown
            fmt_date(p.get("Participant_MED_EffectiveDate")),
            fmt_date(p.get("Participant_DEN_EffectiveDate")),
            fmt_date(p.get("RX_Mailed")), #unkown
            fmt_date(p.get("SPD_Mailed")), #unknown
            fmt_date(p.get("Participant_BLife_EffectiveDate")),
            fmt_date(p.get("Participant_VLife_EffectiveDate")),
            fmt_date(p.get("Participant_MailDate_CERTs")),
        ]
        table_data.append(row)

    col_widths = [
        20,  # #
        50,  # Group
        60,  # Division
        90,  # Participant
        20,  # CL
        20,  # Life Only (orange)
        23,  # Status
        50,  # SwiftMD Mail (light blue)
        50,  # Begin Date
        50,  # Notice Mailed (light blue)
        50,  # Med Eff
        50,  # Dental Eff
        50,  # Rx Mailed (light blue)
        50,  # SPD Mailed (light blue)
        50,  # BLife Eff
        50,  # VLife Eff
        50   # CERTs Mail (light blue)
    ]

    table = Table(table_data, repeatRows=1, colWidths=col_widths)

    style = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),

        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ])

    for r in range(1, len(table_data)):
        style.add('BACKGROUND', (5, r), (5, r), LIGHT_ORANGE)  # Life Only
        style.add('BACKGROUND', (7, r), (7, r), LIGHT_BLUE)    # SwiftMD Mail
        style.add('BACKGROUND', (9, r), (9, r), LIGHT_BLUE)    # Notice Mailed
        style.add('BACKGROUND', (12, r), (12, r), LIGHT_BLUE)  # Rx Mailed
        style.add('BACKGROUND', (13, r), (13, r), LIGHT_BLUE)  # SPD Mailed
        style.add('BACKGROUND', (16, r), (16, r), LIGHT_BLUE)  # CERTs Mail

    table.setStyle(style)
    story.append(table)

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)

    buffer.seek(0)
    return Response(
        buffer.read(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "inline; filename=mailing_letter_06.pdf"}
    )


def generate_mailing_07(participants):
    def fmt_date(dt):
        if not dt:
            return ''
        try:
            return datetime.strptime(str(dt), "%Y-%m-%d").strftime('%m/%d/%Y')
        except ValueError:
            try:
                return datetime.strptime(str(dt), "%Y-%m-%d %H:%M:%S").strftime('%m/%d/%Y')
            except ValueError:
                return str(dt)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.75 * inch
    )

    LIGHT_YELLOW = colors.HexColor("#FFFDE7")
    LIGHT_GREEN  = colors.HexColor("#E8F5E9")

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    normal_style = styles['Normal']

    story = []
    story.append(Paragraph("Voluntary Life Payroll Deduction Request Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 0.2 * inch))

    headers = [
        "", "Group", "Participant", "Begin Date", "End Date", "VL Eff", "Wkly Ded", "Age", "Bracket", "Ded Req Date"
    ]
    table_data = [headers]

    for idx, p in enumerate(participants, start=1):

        age_val = p.get('Participant_Age')
        age = str(age_val).split('.')[0] if age_val is not None else ''

        row = [
            str(idx),
            p.get("Participant_Group_Name", ""),
            p.get("Name", ""),
            fmt_date(p.get("Participant_BeginDate")),
            fmt_date(p.get("Participant_EndDate")),
            fmt_date(p.get("Participant_VLife_EffectiveDate")),
            f"${p['Participant_Weekly_VL_Deduction']:,.2f}" if p.get('Participant_Weekly_VL_Deduction') is not None else '',
            age,
            p.get("Participant_Age_Bracket"),
            fmt_date(p.get("Participant_MailDate_DeductRequest"))
        ]
        table_data.append(row)

    col_widths = [
        20,   # Index
        74,   # Group
        120,  # Participant
        65,   # Begin Date
        65,   # End Date
        65,   # VL Eff
        65,   # Wkly Ded
        35,   # Age
        60,   # Bracket
        80    # Ded Req Date
    ]

    table = Table(table_data, repeatRows=1, colWidths=col_widths)

    style = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),

        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ])

    for r in range(1, len(table_data)):
        # Wkly Ded column (yellow background, blue text, bold)
        style.add('BACKGROUND', (6, r), (6, r), LIGHT_YELLOW)
        style.add('TEXTCOLOR', (6, r), (6, r), colors.blue)
        style.add('FONTNAME', (6, r), (6, r), 'Helvetica-Bold')

        # Ded Req Date column (green background, bold text)
        style.add('BACKGROUND', (9, r), (9, r), LIGHT_GREEN)
        style.add('FONTNAME', (9, r), (9, r), 'Helvetica-Bold')

        # End Date column text red
        style.add('TEXTCOLOR', (4, r), (4, r), colors.red)

    table.setStyle(style)
    story.append(table)

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)

    buffer.seek(0)
    return Response(
        buffer.read(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "inline; filename=vlife_payroll_deduct_req.pdf"}
    )





@reports_billing_bp.route('/replace_tokens', methods=['POST'])
@login_required
def replace_tokens(user):
    import tempfile
    from pdf2docx import Converter
    from docx import Document
    import re
    import mimetypes

    uploaded_file = request.files.get('pdf')
    replacements = request.form.to_dict(flat=False)
    output_format = request.form.get('format', 'pdf').lower()

    if not uploaded_file:
        return jsonify({'error': 'No file uploaded'}), 400

    try:
        file_name = uploaded_file.filename
        content_type = uploaded_file.content_type or mimetypes.guess_type(file_name)[0]
        file_ext = file_name.split('.')[-1].lower()
        values = list(replacements.values())
        replacements_flat = {k: v[0] for k, v in replacements.items()}

        #docx to docx
        if file_ext == "docx":
            docx = Document(uploaded_file)
            for para in docx.paragraphs:
                for key, value in replacements_flat.items():
                    placeholder = f"<<{key}>>"
                    if placeholder in para.text:
                        para.text = para.text.replace(placeholder, value)

            output = io.BytesIO()
            docx.save(output)
            output.seek(0)

            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                as_attachment=True,
                download_name="filled_letter.docx"
            )

        #pdf
        elif file_ext == "pdf":
            import fitz  # PyMuPDF
            pdf_bytes = uploaded_file.read()
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")

            form_filled = False
            value_idx = 0

            for page in doc:
                widgets = page.widgets()
                if widgets:
                    for widget in widgets:
                        if widget.field_type == fitz.PDF_WIDGET_TYPE_TEXT and value_idx < len(values):
                            widget.field_value = values[value_idx][0]
                            widget.update()
                            value_idx += 1
                            form_filled = True

            if not form_filled:
                for page in doc:
                    for token, val_list in replacements.items():
                        value = val_list[0]
                        search_term = f"<<{token}>>"
                        matches = page.search_for(search_term)
                        for rect in matches:
                            page.draw_rect(rect, fill=(1, 1, 1), color=(1, 1, 1))
                            x, y = rect.tl
                            x += 6
                            y += 7
                            page.insert_text((x, y), value, fontsize=9, fontname="helv", color=(0, 0, 0))

            #temp pdf
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
                doc.save(tmp_pdf.name)
                tmp_pdf_path = tmp_pdf.name

            #ability to convert to docx
            if output_format == "docx":
                with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp_docx:
                    docx_path = tmp_docx.name

                cv = Converter(tmp_pdf_path)
                cv.convert(docx_path, start=0, end=None)
                cv.close()

                with open(docx_path, "rb") as f:
                    return send_file(
                        io.BytesIO(f.read()),
                        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        as_attachment=True,
                        download_name="filled_letter.docx"
                    )

            with open(tmp_pdf_path, "rb") as f:
                return send_file(
                    io.BytesIO(f.read()),
                    mimetype="application/pdf",
                    as_attachment=True,
                    download_name="filled_letter.pdf"
                )

        else:
            return jsonify({'error': 'Unsupported file type. Only PDF or DOCX allowed.'}), 400

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
