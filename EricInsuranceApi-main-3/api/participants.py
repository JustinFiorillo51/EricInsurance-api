import datetime
import json
from re import search
import csv
import io
import base64
from flask import Blueprint, Response, current_app, request, jsonify, send_file
from werkzeug.utils import secure_filename
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
from db import Session as DBSession
from utils import build_where_clause, login_required, MIME_TYPES
from models import ParticipantCommunication
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl


participants_bp = Blueprint('participants', __name__)

SORT_DIRECTIONS = {
    'ascend': 'ASC',
    'descend': 'DESC'
}


ParticipantQueryFields = '''
Participant_Pkey, 
Participant_ID, 
Format(Participant_SSN,'000000000') AS Participant_SSN, 
CONCAT(Participant_LastName, ', ', Participant_FirstName, ' ', Participant_MiddleInit) AS [Name],
Participant_LastName,
Participant_MiddleInit,
Participant_FirstName,
Participant_Group_ID,
Participant_Group_Name, 
Participant_Phone_Home, 
Participant_Phone_Cell, 
Participant_EMail,
Participant_Address, 
Participant_Address2,
Participant_City, 
Participant_State, 
Participant_BirthDate, 
Participant_HireDate, 
Participant_BLife_EffectiveDate, 
Participant_BLife_EndDate, 
Participant_TermDate,
Participant_Payments_YN,
Participant_Zip,
Participant_Scheduled_BeginDate,
Participant_Status,
Participant_BeginDate,
Participant_EndDate,
Participant_MED_EffectiveDate,
Participant_MED_EndDate,
Participant_Division,
Participant_Enroll_Form,
Participant_Group_PolicyNum,
Participant_PartTime_YN,
Participant_NonEmployee_YN,
Participant_Spouse_Conditional_YN,
Participant_OptOut_Eligible_YN,

Participant_SwiftMD_MailDate,
Participant_MailDate_Convert_Port,
Participant_MailDate_CERTs,
Participant_MailDate_SummaryPlan,
Participant_MailDate_DeductRequest,
Participant_IDCard_MailDate_Medical,
Participant_IDCard_MailDate_Dental,
Participant_IDCard_MailDate_Vision,
Participant_IDCard_InPacket_YN,
Participant_Eligible_SwiftMD_Mailing_YN,
Participant_Eligible_ConvPort_Mailing_YN,
Participant_Eligible_SRCSIS_Mailing_YN,
Participant_Eligible_SPD_Mailing_YN,
Participant_Eligible_Deduct_Mailing_YN,
Participant_IDCard_Omit_YN,
Participant_Coverage_Notes,
Participant_Packet_Comments,
Participant_Coverage_BasicLife_YN,
Participant_VLife_Waived_YN,
Participant_Coverage_Waived_YN,
Participant_Coverage_GuaranteePay_YN,
Participant_FlexPlan_YN,
Participant_Coverage_MedRX_YN,
Participant_Coverage_Dental_YN,
Participant_Coverage_Vision_YN,
Participant_Coverage_Cobra_YN,
Participant_Coverage_OptOut_YN,
Participant_OptOut_BeginDate,
Participant_OptOut_EndDate,
Participant_91ST_DAY,
Participant_Department,
Participant_Occupation,
Participant_External_ID,
Participant_LIFE_CLASS,
Participant_LIFE_ONLY_YN,
Participant_Dependent_Life_YN,
Participant_Coverage_Type,
Participant_LastUpdated_By,
Participant_LastUpdated_Date,
Participant_BLife_Amount,
Participant_BLife_Count,
Participant_BLife_Prem,
Participant_BLife_ADD_Prem,
Participant_VLife_EffectiveDate,
Participant_VLife_EndDate,
Participant_VLife_EE_Amount,
Participant_VLife_EE_Count,
Participant_VLife_EE_Prem,
Participant_VLife_Spouse_Amount,
Participant_VLife_Spouse_Count,
Participant_VLife_Spouse_Prem,
Participant_VLife_Child_Amount,
Participant_VLife_Child_Count,
Participant_VLife_Child_Prem,
Participant_VLife_Total,
Participant_VLife_Rate,
Participant_VLife_EE_InitialAmount,
Participant_VLife_Spouse_InitialAmount,
Participant_VLife_Child_InitialAmount,
Participant_BLife_Spouse_Amount,
Participant_BLife_Child_Amount,
Participant_BLife_Dep_Units,
Participant_BLife_Dep_Prem,
Participant_BLife_Total,
Participant_BLife_EE_InitialAmount,
Participant_BLife_Spouse_InitialAmount,
Participant_BLife_Child_InitialAmount,
Participant_Next_ReviewDate,
Participant_Weekly_VL_Deduction,
Participant_Monthly_PayrollDeduction,
Participant_Age65_VLReductionDate,
Participant_Age70_VLReductionDate,
Participant_Age75_VLReductionDate,
Participant_Age80_VLReductionDate,
Participant_Age65_BLReductionDate,
Participant_Age70_BLReductionDate,
Participant_Age75_BLReductionDate,
Participant_Age80_BLReductionDate,
Participant_Age,
Participant_Age_Bracket,
Participant_Age_Bracket_ChangeDate,
Participant_Baseline_Date,
Participant_Insurance_Notes,
Participant_Total_PayrollDeductions,
Participant_Total_Adjust,
Participant_Total_Premiums,
Participant_Total_RehireRollover,
Participant_Account_Balance,
Participant_Reconcile_Amount,
Participant_Reconcile_Date,
Participant_PayCode,

[Participant_Total_Guardian_Fwd]
'''


ParticipantQueryFieldsEssential = '''
Participant_Pkey, 
Participant_ID,
Format(Participant_SSN,'000000000') AS Participant_SSN, 
CONCAT(Participant_LastName, ',', Participant_FirstName, ' ', Participant_MiddleInit) AS [Name],
Participant_Group_ID,
Participant_Group_Name, 
Participant_Phone_Home, 
Participant_EMail, 
Participant_Address, 
Participant_City, 
Participant_State, 
Participant_BirthDate, 
Participant_HireDate, 
Participant_BLife_EffectiveDate, 
Participant_TermDate,
Participant_Payments_YN,
Participant_Zip,
Participant_Age,
Participant_Status
'''


# Update a participant
@participants_bp.route('/<string:id>', methods=['PUT'])
@login_required
def update_participant(user, id):
    with DBSession() as session:
        try:

            data = request.get_json()
            if not data:
                return jsonify(), '400 No data provided.'

            sql_set_statment = []

            # Update only provided fields
            for key, value in data.items():
                sql_set_statment.append(f'{key} = :{key}')
            
            sql_set_statment.append('Participant_LastUpdated_By = :Participant_LastUpdated_By')
            sql_set_statment.append('Participant_LastUpdated_Date = :Participant_LastUpdated_Date')

            update_sql = f"UPDATE XParticipant SET {', '.join(sql_set_statment)} WHERE Participant_Pkey = :id"

            session.execute(text(update_sql), {'id': id, **data, 'Participant_LastUpdated_By': user['User_ID'], 'Participant_LastUpdated_Date': datetime.datetime.now(datetime.timezone.utc)})
            session.commit()

            result = session.execute(text(f'''
            SELECT 
                {ParticipantQueryFields}
            FROM dbo.XParticipant
            WHERE XParticipant.Participant_Pkey = :id;
            '''), {'id': id})

            participant = result.first()
            if not participant:
                return jsonify(), '404 Participant not found.'

            return jsonify(participant._asdict()), '200 OK'

        except SQLAlchemyError as e:
            session.rollback()
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/new_participant', methods=['POST'])
@login_required
def new_participant(user):

    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'
    
    if 'Participant_Pkey' not in data:
        return jsonify(), '400 Participant_Pkey is required.'
    
    if 'Participant_ID' not in data:
        return jsonify(), '400 Participant_ID is required.'
    
    if 'Participant_SSN' not in data:
        return jsonify(), '400 Participant_SSN is required.'
    
    if 'Participant_BirthDate' not in data:
        return jsonify(), '400 Participant_BirthDate is required.'
    
    if 'Participant_HireDate' not in data:
        return jsonify(), '400 Participant_HireDate is required.'
    
    try:
        with DBSession() as session:
        
            # Check if the participant already exists
            result = session.execute(text(f'''
            SELECT 1 FROM XParticipant WHERE XParticipant.Participant_Pkey = :Participant_Pkey;
            '''), {'Participant_Pkey': data['Participant_Pkey']})
            if result.scalar():
                return jsonify(), '400 Participant already exists.'

            # Insert an all-null new participant
            insert_sql = f"""
            INSERT INTO XParticipant (
                Participant_Pkey,
                Participant_ID,
                Participant_SSN,
                Participant_BirthDate,
                Participant_HireDate,
                Participant_LastUpdated_By,
                Participant_LastUpdated_Date
            ) VALUES (
                :Participant_Pkey,
                :Participant_ID,
                :Participant_SSN,
                :Participant_BirthDate,
                :Participant_HireDate,
                :Participant_LastUpdated_By,
                :Participant_LastUpdated_Date
            );
            """

            result = session.execute(text(insert_sql), {**data, 'Participant_LastUpdated_By': user['User_ID'], 'Participant_LastUpdated_Date': datetime.datetime.now(datetime.timezone.utc)})
            session.commit()

            return jsonify(), '201 Created'
        
    except Exception as e:
        current_app.logerr(e)
        return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/', methods=['GET'])
@login_required
def participants(user):

    search_by = request.args.get('search_by', 'Name')
    search_by_value = request.args.get('search_by_value', '')
    search_by_group = request.args.get('search_by_group', '')
    search_include_terminated = request.args.get('include_terminated', 'false')
    page = request.args.get('page', 1)
    page_size = request.args.get('page_size', 10)
    sort_by = request.args.get('sort_by', 'Name')
    sort_order = request.args.get('sort_order', 'ascend')

    if search_by not in ('Name', 'Participant_ID', 'Participant_SSN', 'Participant_Phone_Home', 'Participant_Address', 'Participant_Payments_YN'):
        return jsonify(), '400 Parameters are not valid.'
    
    if search_by == 'Participant_Payments_YN':
        if search_by_value not in ('true', 'false'):
            return jsonify(), '400 Parameters are not valid.'
        search_by_value = search_by_value == 'true'
    
    if search_include_terminated not in ('true', 'false'):
        return jsonify(), '400 Parameters are not valid.'
    
    search_include_terminated = search_include_terminated == 'true'

    if isinstance(page, str):
        page = int(page)
    if isinstance(page_size, str):
        page_size = int(page_size)

    sql_params = {
        'offset': (page - 1) * page_size,
        'limit': page_size,
        'search_by_value': search_by_value,
        'terminated': search_include_terminated
    }

    if isinstance(search_by_value, str):
        search_by_clause = f'T1.[{search_by}] LIKE :search_by_value'
        if search_by_value == '':
            search_by_clause = f'{search_by_clause} OR T1.[{search_by}] IS NULL'
        sql_params['search_by_value'] = f'%{search_by_value.replace(' ', '%')}%'
    else:
        search_by_clause = f'T1.[{search_by}] = :search_by_value'

    search_by_terminated_clause = '1=1'
    if not search_include_terminated:
        search_by_terminated_clause = 'T1.Terminated = 0 OR T1.Terminated IS NULL'

    if search_by_group != '':
        search_by_group_clause = f'T1.[Participant_Group_ID] = :search_by_group'
        sql_params['search_by_group'] = search_by_group
    else:
        search_by_group_clause = '1=1'

    if sort_by not in ('Name', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT * FROM (
                SELECT 
                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID
                FROM dbo.vw_participants T1
                    LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
            ) AS T1
            WHERE ({search_by_clause})
            AND ({search_by_group_clause})
            AND ({search_by_terminated_clause})
            ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            OFFSET :offset ROWS
            FETCH NEXT :limit ROWS ONLY;
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()
            data = [dict(zip(columns, row)) for row in rows]

            result_count = session.execute(text(f'''
            SELECT COUNT(1) as [Total] FROM (
                SELECT 
                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID
                FROM dbo.vw_participants T1
                    LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
            ) AS T1
            WHERE ({search_by_clause})
            AND ({search_by_group_clause})
            AND ({search_by_terminated_clause});
            '''), sql_params)
            total = result_count.scalar()

            return jsonify({
                'columns': columns,
                'data': data,
                'total': total
            }), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'

        
        
@participants_bp.route('/export_csv', methods=['GET'])
@login_required
def export_csv(user):

    search_by = request.args.get('search_by', 'Name')
    search_by_value = request.args.get('search_by_value', '')
    search_by_group = request.args.get('search_by_group', '')
    search_include_terminated = request.args.get('include_terminated', 'false')
    sort_by = request.args.get('sort_by', 'Name')
    sort_order = request.args.get('sort_order', 'ascend')

    if search_by not in ('Name', 'Participant_ID', 'Participant_SSN', 'Participant_Phone_Home', 'Participant_Address', 'Participant_Payments_YN'):
        return jsonify(), '400 Parameters are not valid.'
    
    if search_by == 'Participant_Payments_YN':
        if search_by_value not in ('true', 'false'):
            return jsonify(), '400 Parameters are not valid.'
        search_by_value = search_by_value == 'true'
    
    if search_include_terminated not in ('true', 'false'):
        return jsonify(), '400 Parameters are not valid.'
    
    search_include_terminated = search_include_terminated == 'true'

    sql_params = {
        'search_by_value': search_by_value,
        'terminated': search_include_terminated
    }

    if isinstance(search_by_value, str):
        search_by_clause = f'T1.[{search_by}] LIKE :search_by_value'
        if search_by_value == '':
            search_by_clause = f'{search_by_clause} OR T1.[{search_by}] IS NULL'
        sql_params['search_by_value'] = f'%{search_by_value}%'
    else:
        search_by_clause = f'T1.[{search_by}] = :search_by_value'

    search_by_terminated_clause = '1=1'
    if not search_include_terminated:
        search_by_terminated_clause = 'T1.Terminated = 0 OR T1.Terminated IS NULL'

    if search_by_group != '':
        search_by_group_clause = f'T1.[Participant_Group_ID] = :search_by_group'
        sql_params['search_by_group'] = search_by_group
    else:
        search_by_group_clause = '1=1'

    if sort_by == '':
        sort_by = 'Name'
    if sort_order == '':
        sort_order = 'ascend'

    if sort_by not in ('Name', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT * FROM (
                SELECT 
                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID
                FROM dbo.vw_participants T1
                    LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
            ) AS T1
            WHERE ({search_by_clause})
            AND ({search_by_group_clause})
            AND ({search_by_terminated_clause})
            ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()

            # Write CSV to memory
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            for row in rows:
                writer.writerow(list(row))

            output.seek(0)
            csv_data = output.getvalue()
            output.close()

            # Prepare response
            response = Response(
                csv_data,
                mimetype='text/csv',
                headers={
                    "Content-Disposition": f"attachment; filename=participants_{datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%d_%H%M')}.csv"
                }
            )
            return response
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/export_excel', methods=['GET'])
@login_required
def export_excel(user):
    search_by = request.args.get('search_by', 'Name')
    search_by_value = request.args.get('search_by_value', '')
    search_by_group = request.args.get('search_by_group', '')
    search_include_terminated = request.args.get('include_terminated', 'false')
    sort_by = request.args.get('sort_by', 'Name')
    sort_order = request.args.get('sort_order', 'ascend')

    if search_by not in ('Name', 'Participant_ID', 'Participant_SSN', 'Participant_Phone_Home', 'Participant_Address', 'Participant_Payments_YN'):
        return jsonify(), '400 Parameters are not valid.'
    
    if search_by == 'Participant_Payments_YN':
        if search_by_value not in ('true', 'false'):
            return jsonify(), '400 Parameters are not valid.'
        search_by_value = search_by_value == 'true'
    
    if search_include_terminated not in ('true', 'false'):
        return jsonify(), '400 Parameters are not valid.'
    
    search_include_terminated = search_include_terminated == 'true'

    sql_params = {
        'search_by_value': search_by_value,
        'terminated': search_include_terminated
    }

    if isinstance(search_by_value, str):
        search_by_clause = f'T1.[{search_by}] LIKE :search_by_value'
        if search_by_value == '':
            search_by_clause = f'{search_by_clause} OR T1.[{search_by}] IS NULL'
        sql_params['search_by_value'] = f'%{search_by_value}%'
    else:
        search_by_clause = f'T1.[{search_by}] = :search_by_value'

    search_by_terminated_clause = '1=1'
    if not search_include_terminated:
        search_by_terminated_clause = 'T1.Terminated = 0 OR T1.Terminated IS NULL'

    if search_by_group != '':
        search_by_group_clause = f'T1.[Participant_Group_ID] = :search_by_group'
        sql_params['search_by_group'] = search_by_group
    else:
        search_by_group_clause = '1=1'

    if sort_by == '':
        sort_by = 'Name'
    if sort_order == '':
        sort_order = 'ascend'

    if sort_by not in ('Name', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
                SELECT * FROM (
                    SELECT 
                    {ParticipantQueryFields},
                    T1.Terminated,
                    T1.RowID
                    FROM dbo.vw_participants T1
                        LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) 
                        AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
                ) AS T1
                WHERE ({search_by_clause})
                AND ({search_by_group_clause})
                AND ({search_by_terminated_clause})
                ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            '''), sql_params)

            columns = list(result.keys())
            rows = result.fetchall()

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Participants'

            # Write header
            ws.append(columns)

            # Write data
            for row in rows:
                ws.append(list(row))

            # Auto-adjust column widths
            for col_idx, col in enumerate(columns, 1):
                max_length = max((len(str(cell[col_idx - 1])) for cell in rows), default=0)
                ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(40, max_length + 2))

            # Save to in-memory file
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Return Excel file
            return Response(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    "Content-Disposition": f"attachment; filename=participants_{datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
                }
            )

        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'

@participants_bp.route('/search_by_family_members', methods=['GET'])
@login_required
def search_by_family_members(user):

    search_by = request.args.get('search_by', 'Name')
    search_by_value = request.args.get('search_by_value', '')
    search_by_group = request.args.get('search_by_group', '')
    search_include_terminated = request.args.get('include_terminated', 'false')
    page = request.args.get('page', 1)
    page_size = request.args.get('page_size', 10)
    sort_by = request.args.get('sort_by', 'FamilyMemberName')
    sort_order = request.args.get('sort_order', 'ascend')

    if search_by not in ('Name', 'Participant_ID', 'Participant_SSN', 'Participant_Phone_Home', 'Participant_Address', 'Participant_Payments_YN', 'FamilyMemberName'):
        return jsonify(), '400 Parameters are not valid.'

    if search_by == 'Participant_Payments_YN':
        if search_by_value not in ('true', 'false'):
            return jsonify(), '400 Parameters are not valid.'
        search_by_value = search_by_value == 'true'
    
    if search_include_terminated not in ('true', 'false'):
        return jsonify(), '400 Parameters are not valid.'
    
    search_include_terminated = search_include_terminated == 'true'

    if isinstance(page, str):
        page = int(page)
    if isinstance(page_size, str):
        page_size = int(page_size)

    sql_params = {
        'offset': (page - 1) * page_size,
        'limit': page_size,
        'search_by_value': search_by_value,
        'terminated': search_include_terminated
    }

    if isinstance(search_by_value, str):
        search_by_clause = f'T2.[{search_by}] LIKE :search_by_value'
        if search_by_value == '':
            search_by_clause = f'{search_by_clause} OR T2.[{search_by}] IS NULL'
        sql_params['search_by_value'] = f'%{search_by_value}%'
    else:
        search_by_clause = f'T2.[{search_by}] = :search_by_value'

    if search_by_group != '':
        search_by_group_clause = f'T2.[Participant_Group_ID] = :search_by_group'
        sql_params['search_by_group'] = search_by_group
    else:
        search_by_group_clause = '1=1'

    search_by_terminated_clause = '1=1'
    if not search_include_terminated:
        search_by_terminated_clause = 'T2.Terminated = 0 OR T2.Terminated IS NULL'

    if sort_by not in ('Name', 'FamilyMemberName', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT T2.* 
            FROM (
                SELECT 
                CONCAT(XParticipant_Dependents.Depend_Pkey, '_', XParticipant_Dependents.Depend_AutoNum) AS PKey, 
                XParticipant_Dependents.Depend_Pkey AS FamilyMemberPkey, 
                CONCAT(XParticipant_Dependents.Depend_LastName, ',', XParticipant_Dependents.Depend_FirstName, ' ', XParticipant_Dependents.Depend_MiddleName) AS FamilyMemberName, 
                XParticipant_Dependents.Depend_AutoNum as FamilyMemberAutoNum, 
                Format(XParticipant_Dependents.Depend_SSN,'000000000') AS FamilyMemberSSN, 

                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID

                FROM XParticipant_Dependents 
                INNER JOIN vw_participants AS T1 ON XParticipant_Dependents.Depend_Pkey = T1.Participant_Pkey
                WHERE Trim(XParticipant_Dependents.Depend_LastName) > ''
            ) AS T2
            WHERE ({search_by_clause})
            AND ({search_by_group_clause})
            AND ({search_by_terminated_clause})
            ORDER BY T2.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            OFFSET :offset ROWS
            FETCH NEXT :limit ROWS ONLY;
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()
            data = [dict(zip(columns, row)) for row in rows]

            result_count = session.execute(text(f'''
            SELECT COUNT(1) as [Total] 
            FROM (
                SELECT 
                XParticipant_Dependents.Depend_Pkey AS FamilyMemberPkey, 
                CONCAT(XParticipant_Dependents.Depend_LastName, ',', XParticipant_Dependents.Depend_FirstName, ' ', XParticipant_Dependents.Depend_MiddleName) AS FamilyMemberName, 
                XParticipant_Dependents.Depend_AutoNum as FamilyMemberAutoNum, 
                Format(XParticipant_Dependents.Depend_SSN,'000000000') AS FamilyMemberSSN, 

                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID

                FROM XParticipant_Dependents 
                INNER JOIN vw_participants AS T1 ON XParticipant_Dependents.Depend_Pkey = T1.Participant_Pkey
                WHERE Trim(XParticipant_Dependents.Depend_LastName) > ''
            ) AS T2
            WHERE ({search_by_clause})
            AND ({search_by_group_clause})
            AND ({search_by_terminated_clause})
            '''), sql_params)
            total = result_count.scalar()

            return jsonify({
                'columns': columns,
                'data': data,
                'total': total
            }), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'
        

@participants_bp.route('/export_csv_by_family_members', methods=['GET'])
@login_required
def export_csv_by_family_members(user):

    search_by = request.args.get('search_by', 'Name')
    search_by_value = request.args.get('search_by_value', '')
    search_by_group = request.args.get('search_by_group', '')
    search_include_terminated = request.args.get('include_terminated', 'false')
    sort_by = request.args.get('sort_by', 'FamilyMemberName')
    sort_order = request.args.get('sort_order', 'ascend')

    if search_by not in ('Name', 'Participant_ID', 'Participant_SSN', 'Participant_Phone_Home', 'Participant_Address', 'Participant_Payments_YN', 'FamilyMemberName'):
        return jsonify(), '400 Parameters are not valid.'

    if search_by == 'Participant_Payments_YN':
        if search_by_value not in ('true', 'false'):
            return jsonify(), '400 Parameters are not valid.'
        search_by_value = search_by_value == 'true'
    
    if search_include_terminated not in ('true', 'false'):
        return jsonify(), '400 Parameters are not valid.'
    
    search_include_terminated = search_include_terminated == 'true'

    sql_params = {
        'search_by_value': search_by_value,
        'terminated': search_include_terminated
    }

    if isinstance(search_by_value, str):
        search_by_clause = f'T2.[{search_by}] LIKE :search_by_value'
        if search_by_value == '':
            search_by_clause = f'{search_by_clause} OR T2.[{search_by}] IS NULL'
        sql_params['search_by_value'] = f'%{search_by_value}%'
    else:
        search_by_clause = f'T2.[{search_by}] = :search_by_value'

    if search_by_group != '':
        search_by_group_clause = f'T2.[Participant_Group_ID] = :search_by_group'
        sql_params['search_by_group'] = search_by_group
    else:
        search_by_group_clause = '1=1'

    search_by_terminated_clause = '1=1'
    if not search_include_terminated:
        search_by_terminated_clause = 'T2.Terminated = 0 OR T2.Terminated IS NULL'

    if sort_by == '':
        sort_by = 'FamilyMemberName'
    if sort_order == '':
        sort_order = 'ascend'

    if sort_by not in ('Name', 'FamilyMemberName', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT T2.* 
            FROM (
                SELECT 
                CONCAT(XParticipant_Dependents.Depend_Pkey, '_', XParticipant_Dependents.Depend_AutoNum) AS PKey, 
                XParticipant_Dependents.Depend_Pkey AS FamilyMemberPkey, 
                CONCAT(XParticipant_Dependents.Depend_LastName, ',', XParticipant_Dependents.Depend_FirstName, ' ', XParticipant_Dependents.Depend_MiddleName) AS FamilyMemberName, 
                XParticipant_Dependents.Depend_AutoNum as FamilyMemberAutoNum, 
                Format(XParticipant_Dependents.Depend_SSN,'000000000') AS FamilyMemberSSN, 

                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID

                FROM XParticipant_Dependents 
                INNER JOIN vw_participants AS T1 ON XParticipant_Dependents.Depend_Pkey = T1.Participant_Pkey
                WHERE Trim(XParticipant_Dependents.Depend_LastName) > ''
            ) AS T2
            WHERE ({search_by_clause})
            AND ({search_by_group_clause})
            AND ({search_by_terminated_clause})
            ORDER BY T2.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()

            # Write CSV to memory
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            for row in rows:
                writer.writerow(list(row))

            output.seek(0)
            csv_data = output.getvalue()
            output.close()

            # Prepare response
            response = Response(
                csv_data,
                mimetype='text/csv',
                headers={
                    "Content-Disposition": f"attachment; filename=participants_{datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%d_%H%M')}.csv"
                }
            )
            return response
        
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'
    

@participants_bp.route('/<string:id>', methods=['GET'])
@login_required
def get_participant_by_id(user, id):
    with DBSession() as session:
        try:

            result = session.execute(text(f'''
            SELECT 
                {ParticipantQueryFields},
                Terminated,
                RowID
            FROM dbo.vw_participants
            WHERE Participant_Pkey = :id;
            '''), {'id': id})

            participant = result.first()
            if not participant:
                return jsonify(), '404 Participant not found.'
            return jsonify(participant._asdict()), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.' 


@participants_bp.route('/advanced_filter', methods=['POST'])
@login_required
def participants_advanced_filter(user):
    """
    Advanced filter API endpoint that supports complex multi-condition filtering
    """
    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'
    
    filter_criteria = data.get('filter_criteria')
    page = data.get('page', 1)
    page_size = data.get('page_size', 10)
    sort_by = data.get('sort_by', 'Name')
    sort_order = data.get('sort_order', 'ascend')

    if isinstance(page, str):
        page = int(page)
    if isinstance(page_size, str):
        page_size = int(page_size)

    if sort_by not in ('Name', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    sql_params = {
        'offset': (page - 1) * page_size,
        'limit': page_size
    }

    # Build WHERE clause
    where_clause = build_where_clause(filter_criteria, sql_params)

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT * FROM (
                SELECT 
                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID
                FROM dbo.vw_participants T1
                    LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
            ) AS T1
            WHERE {where_clause}
            ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            OFFSET :offset ROWS
            FETCH NEXT :limit ROWS ONLY;
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()
            data = [dict(zip(columns, row)) for row in rows]

            result_count = session.execute(text(f'''
            SELECT COUNT(1) as [Total] FROM (
                SELECT 
                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID
                FROM dbo.vw_participants T1
                    LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
            ) AS T1
            WHERE {where_clause};
            '''), sql_params)
            total = result_count.scalar()

            return jsonify({
                'columns': columns,
                'data': data,
                'total': total
            }), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/by_pkeys', methods=['POST'])
@login_required
def participants_by_pkeys(user):
    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'

    pkeys = data.get('pkeys', [])
    if not pkeys:
        return jsonify(), '400 No pkeys provided.'

    pkeys = [str(pkey) for pkey in pkeys]
    pkeys_placeholders = [f':pkey{i}' for i in range(len(pkeys))]
    params = {f"pkey{i}": id_value for i, id_value in enumerate(pkeys)}

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT 
                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID
            FROM dbo.vw_participants T1
            LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
            WHERE T1.Participant_Pkey IN ({','.join(pkeys_placeholders)});
            '''), params)
            rows = result.all()
            columns = list(result.keys())
            data = [dict(zip(columns, row)) for row in rows]
            return jsonify(data), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/export_csv_advanced', methods=['POST'])
@login_required
def export_csv_advanced(user):
    """
    Advanced filter CSV export
    """
    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'
    
    filter_criteria = data.get('filter_criteria')
    sort_by = data.get('sort_by', 'Name')
    sort_order = data.get('sort_order', 'ascend')

    if sort_by not in ('Name', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    sql_params = {}

    # Build WHERE clause
    where_clause = build_where_clause(filter_criteria, sql_params)

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT * FROM (
                SELECT 
                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID
                FROM dbo.vw_participants T1
                    LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
            ) AS T1
            WHERE {where_clause}
            ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()

            # Write CSV to memory
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            for row in rows:
                writer.writerow(list(row))

            output.seek(0)
            csv_data = output.getvalue()
            output.close()

            # Prepare response
            response = Response(
                csv_data,
                mimetype='text/csv',
                headers={
                    "Content-Disposition": f"attachment; filename=participants_advanced_{datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%d_%H%M')}.csv"
                }
            )
            return response
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'

@participants_bp.route('/export_excel_advanced', methods=['POST'])
@login_required
def export_excel_advanced(user):
    """
    Advanced filter Excel export
    """
    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'
    
    filter_criteria = data.get('filter_criteria')
    sort_by = data.get('sort_by', 'Name')
    sort_order = data.get('sort_order', 'ascend')

    if sort_by not in ('Name', 'Participant_Status', 'Participant_Group_Name'):
        return jsonify(), '400 Parameters are not valid.'
    
    if sort_order not in ('ascend', 'descend'):
        return jsonify(), '400 Parameters are not valid.'

    sql_params = {}

    # Build WHERE clause
    where_clause = build_where_clause(filter_criteria, sql_params)

    with DBSession() as session:
        try:
            result = session.execute(text(f'''
            SELECT * FROM (
                SELECT 
                {ParticipantQueryFields},
                T1.Terminated,
                T1.RowID
                FROM dbo.vw_participants T1
                    LEFT JOIN Table_Rates ON (T1.Participant_Group_PolicyNum = Table_Rates.Rate_PolicyID) AND (T1.Participant_Baseline_Date = Table_Rates.Rate_Date)
            ) AS T1
            WHERE {where_clause}
            ORDER BY T1.[{sort_by}] {SORT_DIRECTIONS[sort_order]}
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()

            # Write Excel to memory
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Participants'

            # Header
            ws.append(columns)

            # Data rows
            for row in rows:
                ws.append(list(row))

            # Save to memory
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Prepare response
            response = Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    "Content-Disposition": f"attachment; filename=participants_advanced_{datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
                }
            )
            return response
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


# Create a communication for a participant
@participants_bp.route('/communications', methods=['POST'])
@login_required
def create_participant_communication(user):
    try:
        data = request.get_json()
        if not data:
            return jsonify(), '400 No data provided.'

        participant_pkey = data.get('Participant_Pkey')
        com_type = data.get('ComType')
        subject = data.get('Subject')
        com_content = data.get('ComContent')
        file_name = data.get('FileName')
        receiver = data.get('Receiver')
        event_date = data.get('EventDate')
        record_status = data.get('RecordStatus', 1)
        image_base64 = data.get('ImageBase64')

        # Validate required fields
        allowed_types = ('SwiftMD', 'ConvPort', 'SRCSIS', 'SPD', 'Deduct', 'MedicalRx', 'Dental', 'Vision')
        if not participant_pkey or not com_type or com_type not in allowed_types:
            return jsonify(), '400 Parameters are not valid.'

        # Normalize/parse values
        if isinstance(record_status, str):
            try:
                record_status = int(record_status)
            except Exception:
                record_status = 1

        parsed_event_date = None
        if event_date:
            try:
                if isinstance(event_date, str):
                    # Accept 'YYYY-MM-DD' or ISO datetime strings
                    try:
                        parsed_event_date = datetime.date.fromisoformat(event_date)
                    except Exception:
                        parsed_event_date = datetime.datetime.fromisoformat(event_date).date()
                elif isinstance(event_date, datetime.date):
                    parsed_event_date = event_date
            except Exception:
                parsed_event_date = None

        image_blob = None
        if image_base64:
            try:
                if isinstance(image_base64, str) and image_base64.startswith('data:'):
                    image_base64 = image_base64.split(',', 1)[1]
                image_blob = base64.b64decode(image_base64)
            except Exception:
                return jsonify(), '400 Parameters are not valid.'

        with DBSession() as session:
            new_record = ParticipantCommunication(
                Participant_Pkey=str(participant_pkey),
                EventDate=parsed_event_date,
                ComType=str(com_type)[:10],
                Subject=(str(subject)[:200] if subject is not None else None),
                ComContent=(str(com_content) if com_content is not None else None),
                ImageBlob=image_blob,
                FileName=(str(file_name)[:300] if file_name is not None else None),
                Receiver=(str(receiver)[:100] if receiver is not None else None),
                CreatedDate=datetime.datetime.now(datetime.timezone.utc),
                CreatedBy=user.get('User_ID') if isinstance(user, dict) else None,
                UpdatedDate=None,
                UpdatedBy=None,
                RecordStatus=record_status,
                DeletedFlag=False
            )

            try:
                session.add(new_record)
                session.flush()  # get RowID
                new_row_id = new_record.RowID
                session.commit()
            except Exception as e:
                session.rollback()
                current_app.logerr(e)
                return jsonify(), '500 An internal error has occurred.'

            return jsonify({'RowID': new_row_id}), '201 Created'
    except Exception as e:
        current_app.logerr(e)
        return jsonify(), '500 An internal error has occurred.'


# Create a communication for a participant via multipart/form-data (file upload)
@participants_bp.route('/communications/upload', methods=['POST'])
@login_required
def create_participant_communication_upload(user):
    try:
        participant_pkey = request.form.get('Participant_Pkey')
        com_type = request.form.get('ComType')
        subject = request.form.get('Subject')
        com_content = request.form.get('ComContent')
        receiver = request.form.get('Receiver')
        event_date = request.form.get('EventDate')
        record_status = request.form.get('RecordStatus', 1)

        allowed_types = ('SwiftMD', 'ConvPort', 'SRCSIS', 'SPD', 'Deduct', 'MedicalRx', 'Dental', 'Vision')
        if not participant_pkey or not com_type or com_type not in allowed_types:
            return jsonify(), '400 Parameters are not valid.'

        try:
            record_status = int(record_status)
        except Exception:
            record_status = 1

        parsed_event_date = None
        if event_date:
            try:
                parsed_event_date = datetime.date.fromisoformat(event_date)
            except Exception:
                try:
                    parsed_event_date = datetime.datetime.fromisoformat(event_date).date()
                except Exception:
                    parsed_event_date = None

        uploaded_file = request.files.get('file')
        image_blob = uploaded_file.read() if uploaded_file else None
        file_name = None
        if uploaded_file and uploaded_file.filename:
            file_name = secure_filename(uploaded_file.filename)[:300]

        with DBSession() as session:
            new_record = ParticipantCommunication(
                Participant_Pkey=str(participant_pkey),
                EventDate=parsed_event_date,
                ComType=str(com_type)[:10],
                Subject=(str(subject)[:200] if subject else None),
                ComContent=(str(com_content) if com_content else None),
                ImageBlob=image_blob,
                FileName=file_name,
                Receiver=(str(receiver)[:100] if receiver else None),
                CreatedDate=datetime.datetime.now(datetime.timezone.utc),
                CreatedBy=user.get('User_ID') if isinstance(user, dict) else None,
                UpdatedDate=None,
                UpdatedBy=None,
                RecordStatus=record_status,
                DeletedFlag=False
            )

            try:
                session.add(new_record)
                session.flush()
                new_row_id = new_record.RowID
                session.commit()
            except Exception as e:
                session.rollback()
                current_app.logerr(e)
                return jsonify(), '500 An internal error has occurred.'

            return jsonify({'RowID': new_row_id}), '201 Created'
    except Exception as e:
        current_app.logerr(e)
        return jsonify(), '500 An internal error has occurred.'


# Communications for a participant (paged, ordered by EventDate DESC)
@participants_bp.route('/communications', methods=['GET'])
@login_required
def get_participant_communications(user):
    participant_pkey = request.args.get('participant_pkey', '')
    page = request.args.get('page', 1)
    page_size = request.args.get('page_size', 5)
    com_type = request.args.get('ComType', '')

    if not participant_pkey or not com_type or com_type not in ('SwiftMD', 'ConvPort', 'SRCSIS', 'SPD', 'Deduct', 'MedicalRx', 'Dental', 'Vision'):
        return jsonify(), '400 Parameters are not valid.'

    if isinstance(page, str):
        page = int(page)
    if isinstance(page_size, str):
        page_size = int(page_size)

    sql_params = {
        'participant_pkey': participant_pkey,
        'ComType': com_type,
        'offset': (page - 1) * page_size,
        'limit': page_size
    }

    with DBSession() as session:
        try:
            result = session.execute(text('''
                SELECT 
                    RowID,
                    Participant_Pkey,
                    EventDate,
                    ComType,
                    Subject,
                    ComContent,
                    FileName,
                    Receiver,
                    CreatedDate,
                    CreatedBy,
                    UpdatedDate,
                    UpdatedBy,
                    RecordStatus
                FROM dbo.XParticipant_Communications
                WHERE Participant_Pkey = :participant_pkey AND ComType = :ComType
                ORDER BY EventDate DESC, RowID DESC
                OFFSET :offset ROWS
                FETCH NEXT :limit ROWS ONLY;
            '''), sql_params)

            columns = list(result.keys())
            rows = result.all()
            data = [dict(zip(columns, row)) for row in rows]

            result_count = session.execute(text('''
                SELECT COUNT(1) as [Total]
                FROM dbo.XParticipant_Communications
                WHERE Participant_Pkey = :participant_pkey AND ComType = :ComType;
            '''), {'participant_pkey': participant_pkey, 'ComType': com_type})
            total = result_count.scalar()

            return jsonify({
                'columns': columns,
                'data': data,
                'total': total
            }), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/communications/<int:row_id>/download', methods=['GET'])
@login_required
def download_participant_communication_file(user, row_id):
    """
    Download the ImageBlob file for a participant communication by RowID
    """
    try:
        with DBSession() as session:
            record = session.query(ParticipantCommunication).filter(ParticipantCommunication.RowID == row_id).first()
            if not record:
                return jsonify({'error': 'Communication record not found'}), 404

            if not record.ImageBlob:
                return jsonify({'error': 'No file content found for this communication'}), 404

            file_stream = io.BytesIO(record.ImageBlob)
            file_stream.seek(0)

            file_name = record.FileName or f"communication_{row_id}"
            file_extension = file_name.split('.')[-1].lower() if '.' in file_name else None
            mime_type = MIME_TYPES.get(file_extension, 'application/octet-stream')

            return send_file(
                file_stream,
                mimetype=mime_type,
                as_attachment=True,
                download_name=file_name
            )
    except Exception as e:
        current_app.logerr(e)
        return jsonify({'error': 'An internal error has occurred'}), 500


# Saved Filters API endpoints
@participants_bp.route('/saved_filters', methods=['GET'])
@login_required
def get_saved_filters(user):
    """
    Get all saved filters for the current user
    """
    with DBSession() as session:
        try:
            result = session.execute(text('''
            SELECT Filter_ID, Filter_Name, Filter_Configuration, User_ID, 
                   Created_Date, LastModified_Date, Is_Active
            FROM dbo.SavedFilters 
            WHERE User_ID = :user_id AND Is_Active = 1
            ORDER BY LastModified_Date DESC
            '''), {'user_id': user['User_ID']})

            filters = []
            for row in result:
                filters.append({
                    'Filter_ID': row.Filter_ID,
                    'Filter_Name': row.Filter_Name,
                    'Filter_Configuration': row.Filter_Configuration,
                    'User_ID': row.User_ID,
                    'Created_Date': row.Created_Date.isoformat() if row.Created_Date else None,
                    'LastModified_Date': row.LastModified_Date.isoformat() if row.LastModified_Date else None,
                    'Is_Active': row.Is_Active
                })

            return jsonify(filters), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/saved_filters', methods=['POST'])
@login_required
def save_filter(user):
    """
    Save a new filter or update existing filter with same name
    """
    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'
    
    filter_name = data.get('name')
    filter_configuration = data.get('groups')
    
    if not filter_name or not filter_configuration:
        return jsonify(), '400 Filter name and configuration are required.'
    
    with DBSession() as session:
        try:
            # Check for existing filter with same name
            existing_filter = session.execute(text('''
            SELECT Filter_ID FROM dbo.SavedFilters 
            WHERE User_ID = :user_id AND Filter_Name = :filter_name
            '''), {'user_id': user['User_ID'], 'filter_name': filter_name})
            
            existing_result = existing_filter.first()
            
            if existing_result:
                # Update existing filter
                session.execute(text('''
                UPDATE dbo.SavedFilters 
                SET Filter_Configuration = :filter_configuration, 
                    LastModified_Date = :last_modified_date,
                    Is_Active = 1
                WHERE Filter_ID = :filter_id
                '''), {
                    'filter_configuration': json.dumps(filter_configuration),
                    'last_modified_date': datetime.datetime.now(datetime.timezone.utc),
                    'filter_id': existing_result.Filter_ID,
                })
                
                session.commit()
                return jsonify({'message': 'Filter updated successfully', 'filter_id': existing_result.Filter_ID}), '200 OK'
            else:
                # Insert new filter
                result = session.execute(text('''
                INSERT INTO dbo.SavedFilters (Filter_Name, Filter_Configuration, User_ID, Created_Date, LastModified_Date, Is_Active)
                VALUES (:filter_name, :filter_configuration, :user_id, :created_date, :last_modified_date, 1)
                '''), {
                    'filter_name': filter_name,
                    'filter_configuration': json.dumps(filter_configuration),
                    'user_id': user['User_ID'],
                    'created_date': datetime.datetime.now(datetime.timezone.utc),
                    'last_modified_date': datetime.datetime.now(datetime.timezone.utc)
                })
                
                session.commit()
                return jsonify({'message': 'Filter saved successfully'}), '201 Created'
        except Exception as e:
            session.rollback()
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/saved_filters/<int:filter_id>', methods=['PUT'])
@login_required
def update_saved_filter(user, filter_id):
    """
    Update an existing filter
    """
    data = request.get_json()
    if not data:
        return jsonify(), '400 No data provided.'
    
    filter_name = data.get('name')
    filter_configuration = data.get('groups')
    
    if not filter_name or not filter_configuration:
        return jsonify(), '400 Filter name and configuration are required.'
    
    with DBSession() as session:
        try:
            # Check if filter exists and belongs to user
            existing = session.execute(text('''
            SELECT Filter_ID FROM dbo.SavedFilters 
            WHERE Filter_ID = :filter_id AND User_ID = :user_id AND Is_Active = 1
            '''), {'filter_id': filter_id, 'user_id': user['User_ID']})
            
            if not existing.scalar():
                return jsonify(), '404 Filter not found.'
            
            # Check if new name conflicts with other filters
            name_conflict = session.execute(text('''
            SELECT 1 FROM dbo.SavedFilters 
            WHERE User_ID = :user_id AND Filter_Name = :filter_name AND Filter_ID != :filter_id AND Is_Active = 1
            '''), {'user_id': user['User_ID'], 'filter_name': filter_name, 'filter_id': filter_id})
            
            if name_conflict.scalar():
                return jsonify({'error': 'Filter name already exists'}), '400 Filter name already exists.'
            
            # Update filter
            session.execute(text('''
            UPDATE dbo.SavedFilters 
            SET Filter_Name = :filter_name, 
                Filter_Configuration = :filter_configuration, 
                LastModified_Date = :last_modified_date
            WHERE Filter_ID = :filter_id AND User_ID = :user_id
            '''), {
                'filter_name': filter_name,
                'filter_configuration': json.dumps(filter_configuration),
                'last_modified_date': datetime.datetime.now(datetime.timezone.utc),
                'filter_id': filter_id,
                'user_id': user['User_ID']
            })
            
            session.commit()
            
            return jsonify({'message': 'Filter updated successfully'}), '200 OK'
        except Exception as e:
            session.rollback()
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/saved_filters/<int:filter_id>', methods=['DELETE'])
@login_required
def delete_saved_filter(user, filter_id):
    """
    Delete a saved filter (soft delete)
    """
    with DBSession() as session:
        try:
            # Check if filter exists and belongs to user
            existing = session.execute(text('''
            SELECT Filter_ID FROM dbo.SavedFilters 
            WHERE Filter_ID = :filter_id AND User_ID = :user_id AND Is_Active = 1
            '''), {'filter_id': filter_id, 'user_id': user['User_ID']})
            
            if not existing.scalar():
                return jsonify(), '404 Filter not found.'
            
            # Soft delete
            session.execute(text('''
            UPDATE dbo.SavedFilters 
            SET Is_Active = 0, LastModified_Date = :last_modified_date
            WHERE Filter_ID = :filter_id AND User_ID = :user_id
            '''), {
                'last_modified_date': datetime.datetime.now(datetime.timezone.utc),
                'filter_id': filter_id,
                'user_id': user['User_ID']
            })
            
            session.commit()
            
            return jsonify({'message': 'Filter deleted successfully'}), '200 OK'
        except Exception as e:
            session.rollback()
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'


@participants_bp.route('/saved_filters/<int:filter_id>', methods=['GET'])
@login_required
def get_saved_filter_by_id(user, filter_id):
    """
    Get a specific saved filter by ID
    """
    with DBSession() as session:
        try:
            result = session.execute(text('''
            SELECT Filter_ID, Filter_Name, Filter_Configuration, User_ID, 
                   Created_Date, LastModified_Date, Is_Active
            FROM dbo.SavedFilters 
            WHERE Filter_ID = :filter_id AND User_ID = :user_id AND Is_Active = 1
            '''), {'filter_id': filter_id, 'user_id': user['User_ID']})

            filter_data = result.first()
            if not filter_data:
                return jsonify(), '404 Filter not found.'
            
            return jsonify({
                'Filter_ID': filter_data.Filter_ID,
                'Filter_Name': filter_data.Filter_Name,
                'Filter_Configuration': filter_data.Filter_Configuration,
                'User_ID': filter_data.User_ID,
                'Created_Date': filter_data.Created_Date.isoformat() if filter_data.Created_Date else None,
                'LastModified_Date': filter_data.LastModified_Date.isoformat() if filter_data.LastModified_Date else None,
                'Is_Active': filter_data.Is_Active
            }), '200 OK'
        except Exception as e:
            current_app.logerr(e)
            return jsonify(), '500 An internal error has occurred.'

